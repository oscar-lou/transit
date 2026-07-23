"""Tests for _heal_headers() - recovers a header cell corrupted at the
source (a real CrowdStrike workstation export once had a column's name
replaced with the literal number 0).
"""
import openpyxl
import pytest

import consolidate_noncompliant as cnc

EXPECTED = ["gis_bu", "hostname", "install_status", "os", "last_seen", "agent_version",
            "proc_agent_installed", "proc_cs_version_status", "proc_agent_reporting",
            "Compliance", "report_date"]


@pytest.mark.parametrize("value,expected", [
    ("0", True), ("123", True), ("-1", True), ("3.14", True), ("", True), (None, True),
    ("proc_cs_version_status", False), ("b_v2", False), ("0x", False), ("a1", False),
])
def test_looks_like_corruption(value, expected):
    assert cnc._looks_like_corruption(value) is expected


def test_heals_the_real_world_corrupted_column():
    headers = list(EXPECTED)
    headers[7] = "0"  # the actual corruption seen in production
    assert cnc._heal_headers(headers, EXPECTED) == EXPECTED


def test_does_not_touch_a_clean_header_row():
    headers = list(EXPECTED)
    assert cnc._heal_headers(headers, EXPECTED) == headers


def test_does_not_clobber_a_genuine_two_column_reorder():
    """If two expected columns are simply swapped, both names are still
    present elsewhere in the row - the guard must leave a real reorder
    alone rather than 'fixing' it back to the original order."""
    expected = ["a", "b", "c"]
    headers = ["a", "c", "b"]
    assert cnc._heal_headers(headers, expected) == headers


def test_skips_when_column_count_differs():
    expected = ["a", "b", "c"]
    headers = ["a", "b"]
    assert cnc._heal_headers(headers, expected) == headers


def test_skips_when_no_expected_headers_given():
    headers = ["x", "y"]
    assert cnc._heal_headers(headers, None) == headers
    assert cnc._heal_headers(headers, []) == headers


def test_genuine_novel_rename_is_left_alone():
    """Regression for a real, previously-accepted trade-off: a value that
    still LOOKS like a plausible column name (has letters, e.g. a deliberate
    upstream rename 'b' -> 'b_v2') must NOT be healed back to the old name -
    that would silently mask the fact anything changed. Only genuinely
    corruption-shaped values (blank, or purely numeric - see
    _looks_like_corruption) get healed; see
    test_heals_the_real_world_corrupted_column and
    test_heals_a_blank_corrupted_column below for what still does."""
    expected = ["a", "b", "c"]
    headers = ["a", "b_v2", "c"]
    assert cnc._heal_headers(headers, expected) == headers, (
        "REGRESSION: a plausible rename got silently rewritten back to the old name")


def test_heals_a_blank_corrupted_column():
    """Blank is corruption-shaped too (no letters, can't be a real column
    name), same as the literal-number-0 case already covered above."""
    expected = ["a", "b", "c"]
    headers = ["a", "", "c"]
    assert cnc._heal_headers(headers, expected) == expected


def test_corrupted_header_end_to_end_still_classifies_correctly(tmp_path):
    """Drives a genuinely corrupted xlsx (a header cell replaced with the
    literal number 0, exactly as seen in production for this file family)
    through the FULL read path - _read_xlsx_rows -> normalize_file ->
    cs_issue - and asserts the resulting finding is classified correctly,
    not blank. The heal function itself is already pinned in isolation
    above; this pins the downstream result, so a change that breaks the
    wiring between healing and classification (not just the heal itself)
    would be caught here.

    Uses FILE_REGISTRY's own column list (not a local constant) so this
    test can't silently drift from what normalize_file() actually expects.
    Corrupts the 'crowdstrike' column specifically - the one actually
    mapped to agent_installed and driving classification for this source
    (see the "crowdstrike" FILE_REGISTRY entry).
    """
    columns = cnc.FILE_REGISTRY["crowdstrike"]["columns"]
    corrupted_headers = list(columns)
    corrupt_at = columns.index("crowdstrike")
    corrupted_headers[corrupt_at] = 0  # literal int, matching the real corruption openpyxl reads back

    values_by_column = {
        "hostname": "HHOWKLC-TEST01", "business_unit_code": "AIAGO", "manufacturer": "Lenovo",
        "chassis_type": "Notebook", "model_id": "X1", "serial_number": "SN1",
        "company": "AIA Group Office", "assigned_to": "Test User", "hardware_status": "Installed",
        "install_status": "Installed", "os": "Windows 11 Enterprise", "os_domain": "AIA.BIZ",
        "u_vlan": "", "u_dr_availability": "", "u_dr_grouping": "", "u_security_zone": "",
        "sys_class_name": "Computer", "last_discovered": "2026-07-01 00:00:00",
        "business_unit": "", "virtual": "0", "u_non_discoverable_ci": "",
        "crowdstrike_required": "Yes", "crowdstrike": "No", "agent_version": "",
        "report_date": "2026-07-01", "crowdstrike_installed": "0", "host_name": "HHOWKLC-TEST01",
        "compliant": "0", "ageing_30_days": "0", "ageing_60_days": "0", "ageing_90_days": "0",
        "last_seen": "2026-07-01 00:00:00", "run_at": "2026-07-01",
    }
    data_row = [values_by_column[c] for c in columns]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(corrupted_headers)
    ws.append(data_row)
    path = tmp_path / "Crowdstrike_Deployment.xlsx"
    wb.save(path)

    rows = cnc.normalize_file(str(path), registry_key="crowdstrike")

    assert len(rows) == 1, f"expected exactly one normalized row, got {rows!r}"
    assert rows[0]["issue"] == "CrowdStrike agent not installed", (
        f"REGRESSION: header self-heal did not propagate to correct "
        f"classification - got issue={rows[0]['issue']!r}")


def test_read_headers_does_not_iterate_past_the_header_row(tmp_path, monkeypatch):
    """Direct, unambiguous proof (not inferred from timing) that
    _read_headers() still stops after row 1 now that its xlsx-reading logic
    is shared with _read_xlsx_rows() via _open_xlsx_headers() - guards
    against a future refactor accidentally making header-only reads consume
    the whole sheet again. Instruments the real row iterator to COUNT pulls,
    rather than trusting that 'no code after the return happens to touch
    it' - if _read_headers() were ever changed to (even accidentally)
    iterate further, this fails with an exact count, not a vague timing
    difference."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["col1", "col2"])
    for i in range(500):
        ws.append([f"val{i}a", f"val{i}b"])
    path = tmp_path / "big.xlsx"
    wb.save(path)

    real_load_workbook = openpyxl.load_workbook
    pulled = {"count": 0}

    def counting_load_workbook(*args, **kwargs):
        real_wb = real_load_workbook(*args, **kwargs)
        real_ws = real_wb.active
        real_iter_rows = real_ws.iter_rows

        def counting_iter_rows(*a, **k):
            for row in real_iter_rows(*a, **k):
                pulled["count"] += 1
                yield row

        real_ws.iter_rows = counting_iter_rows
        return real_wb

    monkeypatch.setattr(cnc.openpyxl, "load_workbook", counting_load_workbook)
    headers = cnc._read_headers(str(path))

    assert headers == ["col1", "col2"]
    assert pulled["count"] == 1, (
        f"REGRESSION: _read_headers() pulled {pulled['count']} row(s) from a "
        f"501-row sheet, expected exactly 1 (the header row) - it's reading "
        f"past the header again.")
