"""Tests for the xlsx workbook cache (_get_cached_workbook()/
_clear_xlsx_cache()) - added because a large multi-tab workbook like
CompliantReport(Working).xlsx was being opened by openpyxl once per lookup
(once for the pre-flight header check, once each for CMDB/AD_Users/
Overrides), even though every lookup in one run wants the exact same file.

Direct, unambiguous proof (not inferred from timing) that openpyxl.
load_workbook() is called exactly once - the same "prove it, don't just
claim it" standard as test_header_healing.py's row-iterator-count test.
"""
import openpyxl

import consolidate_noncompliant as cnc

# data_dir fixture lives in conftest.py; the autouse _clear_xlsx_workbook_cache
# fixture there also guarantees this file's tests never see a workbook
# cached by a previous test.


def _write_compliant_report(path) -> None:
    """A minimal multi-tab workbook shaped like the real
    CompliantReport(Working).xlsx: a CMDB tab (hostname -> assigned name)
    and an AD_Users tab (name -> email), plus an unrelated third tab (the
    real file has ~20 monthly-snapshot tabs nothing here reads)."""
    wb = openpyxl.Workbook()
    wb.active.title = "SomeOtherTab"
    wb.active.append(["irrelevant"])

    cmdb = wb.create_sheet("CMDB")
    cmdb.append(["Name", "Assigned to"])
    cmdb.append(["WS-1", "Doe, Jane"])

    ad = wb.create_sheet("AD_Users")
    ad.append(["DisplayName", "EmailAddress"])
    ad.append(["Doe, Jane", "jane.doe@example.com"])

    wb.save(path)


def _count_load_workbook_calls(monkeypatch):
    real_load_workbook = openpyxl.load_workbook
    calls = {"count": 0}

    def counting_load_workbook(*args, **kwargs):
        calls["count"] += 1
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(cnc.openpyxl, "load_workbook", counting_load_workbook)
    return calls


def test_multi_tab_workbook_opened_only_once_per_run(data_dir, monkeypatch):
    """The exact real scenario: CMDB lookup, AD_Users lookup, and the
    pre-flight header check all need the SAME workbook. openpyxl.
    load_workbook() must be called exactly once across all three, not once
    per caller - each of these three entry points calls find_dataset()
    (which itself scans every .xlsx's sheet list) and/or _read_headers()/
    _read_xlsx_rows() independently, so this also proves the cache survives
    across multiple different top-level callers, not just repeated calls to
    the same function."""
    _write_compliant_report(data_dir / "CompliantReport(Working).xlsx")
    calls = _count_load_workbook_calls(monkeypatch)

    cnc.validate_headers()
    cnc.read_cmdb_mapping()
    cnc.read_ad_users()

    assert calls["count"] == 1, (
        f"REGRESSION: workbook opened {calls['count']} time(s) in one run, expected exactly 1")


def test_cache_is_cleared_between_runs(data_dir, monkeypatch):
    """The other half of the requirement: a fresh run (main()/build_groups(),
    simulated here by explicitly calling _clear_xlsx_cache()) must NOT reuse
    a workbook cached by an earlier one - two consecutive runs each open the
    file once, for a total of two opens, not one."""
    _write_compliant_report(data_dir / "CompliantReport(Working).xlsx")
    calls = _count_load_workbook_calls(monkeypatch)

    cnc.read_cmdb_mapping()
    assert calls["count"] == 1

    cnc._clear_xlsx_cache()
    cnc.read_cmdb_mapping()
    assert calls["count"] == 2, (
        "REGRESSION: second run reused the first run's cached workbook instead of re-reading")


def test_cached_workbook_returns_correct_data_for_every_caller(data_dir):
    """Not just 'opened once' - the CACHED workbook must still return the
    right sheet/rows to every distinct caller, proving iter_rows() on a
    shared, reused read-only worksheet genuinely re-reads from the start
    each time rather than returning stale or exhausted results."""
    _write_compliant_report(data_dir / "CompliantReport(Working).xlsx")

    cmdb_names = cnc.read_cmdb_mapping()
    ad = cnc.read_ad_users()

    assert cmdb_names == {"WS-1": "Doe, Jane"}
    assert ad["exact"][cnc.norm_name("Doe, Jane")] == "jane.doe@example.com"

    # Calling both again (same cached workbook) must return the identical,
    # still-correct result - not empty/exhausted from the first read.
    assert cnc.read_cmdb_mapping() == cmdb_names
    assert cnc.read_ad_users()["exact"][cnc.norm_name("Doe, Jane")] == "jane.doe@example.com"


def test_cache_keyed_by_data_dir_not_just_filename(tmp_path, monkeypatch):
    """Two different DATA_DIRs with a file of the SAME name must never share
    a cache entry - otherwise a test (or a future caller) redirecting
    DATA_DIR mid-process could serve one directory's content for another's
    file of the same name."""
    dir_a = tmp_path / "a"
    dir_b = tmp_path / "b"
    dir_a.mkdir()
    dir_b.mkdir()

    wb_a = openpyxl.Workbook()
    wb_a.active.append(["Name", "Assigned to"])
    wb_a.active.append(["HOST-A", "Person A"])
    wb_a.save(dir_a / "CMDB_Mapping.xlsx")

    wb_b = openpyxl.Workbook()
    wb_b.active.append(["Name", "Assigned to"])
    wb_b.active.append(["HOST-B", "Person B"])
    wb_b.save(dir_b / "CMDB_Mapping.xlsx")

    monkeypatch.setattr(cnc, "DATA_DIR", str(dir_a))
    result_a = cnc.read_cmdb_mapping()

    monkeypatch.setattr(cnc, "DATA_DIR", str(dir_b))
    result_b = cnc.read_cmdb_mapping()

    assert result_a == {"HOST-A": "Person A"}
    assert result_b == {"HOST-B": "Person B"}, (
        "REGRESSION: dir_b's read returned dir_a's cached content - cache key must include DATA_DIR")
