"""Tests for find_dataset()'s filename/sheet-tab matching. Includes tests
that PIN a real, demonstrated limitation (substring matching can false-
positive on an unrelated file, or silently miss a renamed one) rather than
hiding it - so a future fix to the matching strategy is a deliberate,
visible change to these tests, not a silent behavior shift. What IS fixed:
when more than one file/sheet matches the same keyword, the result is now
deterministic (sorted, not raw os.listdir() order) and printed as a warning
naming every candidate - see test_ambiguous_filename_match_is_deterministic
and test_ambiguous_match_prints_a_warning.
"""
import os

import openpyxl
import pytest

import consolidate_noncompliant as cnc

# data_dir fixture lives in conftest.py (shared with test_data_source.py)


def _write_blank_xlsx(path) -> None:
    """A real (if empty) xlsx file - needed even for filename-only match
    tests, because find_dataset() falls back to actually opening every
    .xlsx in data/ with openpyxl when no filename matches, so a fake/empty
    byte string raises BadZipFile instead of just failing to match."""
    openpyxl.Workbook().save(path)


def test_bitlocker_registry_key_matches_real_filename(data_dir):
    """The real BitLocker export is named
    '20260715-AIAGO-19. Hard Disk Encryption Compliance.csv' - it never
    contains the word 'bitlocker' anywhere, which is exactly why
    FILE_REGISTRY's key for it is 'encryption', not 'bitlocker' (see the
    comment on that registry entry). Pins that the chosen key actually
    matches the real filename, not just a hypothetical one."""
    (data_dir / "20260715-AIAGO-19. Hard Disk Encryption Compliance.csv").write_text("h1,h2\n")
    path, sheet = cnc.find_dataset("encryption")
    assert path is not None, "REGRESSION: registry key no longer matches the real filename"
    assert "Hard Disk Encryption Compliance" in path
    assert sheet is None


def test_matches_by_filename_substring(data_dir):
    _write_blank_xlsx(data_dir / "AIAGO_Workstation_CS.xlsx")
    path, sheet = cnc.find_dataset("aiago_workstation_cs")
    assert path.endswith("AIAGO_Workstation_CS.xlsx")
    assert sheet is None


def test_returns_none_when_nothing_matches(data_dir):
    _write_blank_xlsx(data_dir / "Totally_Unrelated_File.xlsx")
    assert cnc.find_dataset("aiago_workstation_cs") is None


def test_renaming_a_known_file_makes_it_invisible(data_dir):
    """A real risk: renaming AIAGO_Workstation_CS.xlsx to something that no
    longer contains 'aiago_workstation_cs' as a substring makes the whole
    report vanish from the pipeline with no error - just a console
    '! not found' line in load_all(). This test documents that behavior."""
    _write_blank_xlsx(data_dir / "AIAGO_Workstation_CrowdStrike_Report.xlsx")
    assert cnc.find_dataset("aiago_workstation_cs") is None


def test_false_positive_on_coincidental_filename_substring(data_dir):
    """A real, demonstrated risk: short registry keys like 'dlp' match ANY
    filename containing that substring, even an unrelated file."""
    _write_blank_xlsx(data_dir / "Random_DLP_Meeting_Notes.xlsx")
    found = cnc.find_dataset("dlp")
    assert found is not None
    assert "Random_DLP_Meeting_Notes" in found[0]


def test_ambiguous_filename_match_is_deterministic(data_dir):
    """When two files both match the same keyword, the result must be
    reproducible (sorted order) rather than whatever os.listdir() happens to
    return, which is filesystem/OS-dependent and not guaranteed stable."""
    _write_blank_xlsx(data_dir / "Zeta_DLP_Extra.xlsx")
    _write_blank_xlsx(data_dir / "Alpha_DLP_Extra.xlsx")

    path1, _ = cnc.find_dataset("dlp")
    path2, _ = cnc.find_dataset("dlp")

    assert os.path.basename(path1) == "Alpha_DLP_Extra.xlsx", (
        f"expected the alphabetically-first match, got {os.path.basename(path1)!r}")
    assert path1 == path2, "REGRESSION: repeated calls resolved to different files"


# ===========================================================================
# Dated-file tie-break - a real, now-observed case: a newer weekly export
# (e.g. 20260715) can land in data/ before an older one (e.g. 20260703) of
# the same report is cleaned out. Plain alphabetically-first (the rule
# above, for UNDATED ambiguous matches) would silently pick the OLDER file,
# since ISO-style YYYYMMDD prefixes sort ascending - exactly the wrong
# direction. See _pick_latest_dated().
# ===========================================================================

def test_ambiguous_dated_files_resolves_to_the_newer_export(data_dir, capsys):
    """The exact real scenario this exists for: two dated copies of the same
    DLP export sitting in data/ together. Must resolve to the NEWER one, not
    the alphabetically-first (= chronologically OLDEST, for ascending
    YYYYMMDD prefixes) one the plain ambiguity rule would otherwise pick."""
    _write_blank_xlsx(data_dir / "20260703_AIAGO-17. Workstation Security Agent Deployment-DLP.xlsx")
    _write_blank_xlsx(data_dir / "20260715_AIAGO-17. Workstation Security Agent Deployment-DLP.xlsx")

    path, sheet = cnc.find_dataset("dlp")

    assert "20260715" in path, f"REGRESSION: expected the newer (20260715) export, got {path!r}"
    assert sheet is None
    out = capsys.readouterr().out
    assert "latest dated" in out
    assert "20260703" in out and "20260715" in out


def test_pick_latest_dated_is_order_independent():
    """Proves the selection is a genuine max-by-parsed-date, not "whichever
    happens to be last in the list" - both orderings of the same two
    candidates must resolve to the same (newer) one."""
    older = "20260703_AIAGO-17. Workstation Security Agent Deployment-DLP.csv"
    newer = "20260715_AIAGO-17. Workstation Security Agent Deployment-DLP.csv"

    assert cnc._pick_latest_dated([older, newer]) == newer
    assert cnc._pick_latest_dated([newer, older]) == newer


def test_pick_latest_dated_prefers_any_dated_file_over_undated():
    """A dated candidate must win over an undated one regardless of
    alphabetical position - '0_Backup...' sorts before the dated file here
    (digits '0' < '2' in ASCII), so this specifically rules out "just take
    whichever is alphabetically first among the non-dated-looking ones"."""
    undated_but_alphabetically_first = "0_Backup_DLP_Extra.csv"
    dated = "20260715_AIAGO-17. Workstation Security Agent Deployment-DLP.csv"

    assert cnc._pick_latest_dated([undated_but_alphabetically_first, dated]) == dated


def test_pick_latest_dated_falls_back_to_first_when_none_dated():
    """No candidate carries a recognizable date prefix -> falls back to the
    existing, already-tested alphabetically-first behavior, unchanged."""
    matches = ["Alpha_DLP_Extra.xlsx", "Beta_DLP_Extra.xlsx"]
    assert cnc._pick_latest_dated(matches) == "Alpha_DLP_Extra.xlsx"


def test_ambiguous_match_prints_a_warning(data_dir, capsys):
    """More than one candidate for the same keyword must be surfaced, not
    silently resolved - a human should notice two reports (or an unrelated
    file) are colliding on the same registry key."""
    _write_blank_xlsx(data_dir / "Alpha_DLP_Extra.xlsx")
    _write_blank_xlsx(data_dir / "Beta_DLP_Extra.xlsx")

    cnc.find_dataset("dlp")

    out = capsys.readouterr().out
    assert "matches 2 files" in out
    assert "Alpha_DLP_Extra.xlsx" in out
    assert "Beta_DLP_Extra.xlsx" in out


def test_single_match_prints_no_ambiguity_warning(data_dir, capsys):
    _write_blank_xlsx(data_dir / "AIAGO_Workstation_CS.xlsx")
    cnc.find_dataset("aiago_workstation_cs")
    out = capsys.readouterr().out
    assert "matches" not in out


def test_matches_a_sheet_inside_a_multi_tab_workbook(data_dir):
    """When no filename matches, find_dataset() falls back to checking sheet
    names inside every .xlsx in data/ (e.g. CompliantReport(Working).xlsx's
    'CMDB' and 'AD_Users' tabs)."""
    wb = openpyxl.Workbook()
    wb.active.title = "26May"
    wb.create_sheet("CMDB")
    wb.create_sheet("AD_Users")
    wb.save(data_dir / "CompliantReport(Working).xlsx")

    path, sheet = cnc.find_dataset("cmdb")
    assert sheet == "CMDB"

    path, sheet = cnc.find_dataset("ad_users")
    assert sheet == "AD_Users"


def test_standalone_file_takes_priority_over_a_sheet_match(data_dir):
    """load_all()'s documented priority: a standalone file wins over a sheet
    living inside some other workbook, checked first (filename pass happens
    before the sheet-tab pass)."""
    (data_dir / "CMDB_Mapping.xlsx").write_bytes(b"")
    wb = openpyxl.Workbook()
    wb.create_sheet("CMDB")
    wb.save(data_dir / "Other_Workbook.xlsx")

    path, sheet = cnc.find_dataset("cmdb")
    assert path.endswith("CMDB_Mapping.xlsx")
    assert sheet is None


def test_require_openpyxl_raises_clear_error_when_missing(monkeypatch):
    """Regression: missing openpyxl used to surface as a bare NameError deep
    in a stack trace (e.g. in a locked-down runtime that doesn't have it
    installed) instead of a message saying what to install."""
    monkeypatch.setattr(cnc, "HAVE_XLSX", False)
    with pytest.raises(RuntimeError, match="openpyxl is required"):
        cnc._require_openpyxl("some_file.xlsx")


def test_require_openpyxl_is_a_noop_when_available(monkeypatch):
    monkeypatch.setattr(cnc, "HAVE_XLSX", True)
    cnc._require_openpyxl("some_file.xlsx")  # must not raise


def test_read_xlsx_rows_raises_clear_error_not_nameerror_when_openpyxl_missing(
        data_dir, monkeypatch):
    path = data_dir / "AIAGO_Workstation_CS.xlsx"
    openpyxl.Workbook().save(path)
    monkeypatch.setattr(cnc, "HAVE_XLSX", False)
    with pytest.raises(RuntimeError, match="openpyxl is required"):
        cnc._read_xlsx_rows(str(path))


def test_list_sheets_raises_clear_error_not_nameerror_when_openpyxl_missing(
        data_dir, monkeypatch):
    path = data_dir / "Some_Workbook.xlsx"
    openpyxl.Workbook().save(path)
    monkeypatch.setattr(cnc, "HAVE_XLSX", False)
    with pytest.raises(RuntimeError, match="openpyxl is required"):
        cnc._list_sheets(str(path))


# ===========================================================================
# Filename attribution on genuine read failures (corrupted .xlsx, bad-encoding
# .csv) - distinct from the missing-openpyxl tests above: those fail before
# ever touching file bytes, these fail while actually reading them, and
# previously gave zero indication of which file was at fault (confirmed:
# 'BadZipFile: File is not a zip file' / 'UnicodeDecodeError: ... invalid
# start byte', neither naming a source, in a directory that can hold 7+
# input files).
# ===========================================================================

def test_read_xlsx_rows_raises_with_filename_when_corrupted(data_dir):
    (data_dir / "AIAGO_Workstation_CS.xlsx").write_bytes(b"not a real xlsx file")
    with pytest.raises(Exception) as excinfo:
        cnc._read_xlsx_rows("AIAGO_Workstation_CS.xlsx")
    assert "AIAGO_Workstation_CS.xlsx" in str(excinfo.value)


def test_read_headers_raises_with_filename_when_xlsx_corrupted(data_dir):
    (data_dir / "AIAGO_Workstation_CS.xlsx").write_bytes(b"not a real xlsx file")
    with pytest.raises(Exception) as excinfo:
        cnc._read_headers("AIAGO_Workstation_CS.xlsx")
    assert "AIAGO_Workstation_CS.xlsx" in str(excinfo.value)


def test_read_csv_rows_raises_with_filename_on_bad_encoding(data_dir):
    (data_dir / "Overrides.csv").write_bytes(b"\xff\xfe\x00\x01bad bytes")
    with pytest.raises(Exception) as excinfo:
        cnc._read_csv_rows("Overrides.csv")
    assert "Overrides.csv" in str(excinfo.value)


def test_read_headers_raises_with_filename_on_bad_encoding(data_dir):
    (data_dir / "Overrides.csv").write_bytes(b"\xff\xfe\x00\x01bad bytes")
    with pytest.raises(Exception) as excinfo:
        cnc._read_headers("Overrides.csv")
    assert "Overrides.csv" in str(excinfo.value)


def test_list_sheets_raises_with_filename_when_corrupted(data_dir):
    (data_dir / "CompliantReport(Working).xlsx").write_bytes(b"not a real xlsx file")
    with pytest.raises(Exception) as excinfo:
        cnc._list_sheets("CompliantReport(Working).xlsx")
    assert "CompliantReport(Working).xlsx" in str(excinfo.value)
