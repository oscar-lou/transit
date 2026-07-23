#!/usr/bin/env python3
"""
Non-compliant report -> per-recipient notification builder.

Pipeline:
  1. Read the several attached .xlsx/.csv reports (CrowdStrike + Purview +
     Zapp + DLP, across workstation / server / Mac), which use different
     column names for the same things, and reconcile them into ONE canonical
     worklist.
  2. Resolve WHO to notify for each finding:
       - workstation: assigned_to (Purview) -> else CMDB hostname->email
                      lookup, via AD_Users (name -> email, fuzzy) and
                      Overrides (manual, authoritative)
       - server:      not notified - servers still appear in the worklist for
                      visibility, but build_notifications() skips them entirely
  3. Consolidate per recipient (one message per person, not one per file) and
     compose an Outlook email + a Teams message.
  4. STUB the send from this script: write a preview workbook of exactly what
     would go out - nothing here ever sends. Real Outlook sending is a
     separate, explicit step: see send_email.py, which already implements it
     in full (dry-run / send-to-self / send-live, with guardrails and an
     audit log) via Microsoft Graph - the only thing still pending is the
     external Mail.Send admin-consent grant, not any missing code. Teams
     sending was never implemented anywhere: compose_teams() below only
     feeds the preview's 'Teams Message' column, not an actual send path.

Inputs (drop into data/, .xlsx or .csv):
  - the 5 core report files (AIAGO_*_CS.xlsx / AIAGO_*_Purview.xlsx), plus
    the DLP and Zapp exports (fuller/additional coverage - see FILE_REGISTRY)
  - a CMDB export named like 'CMDB_Mapping.xlsx' (hostname -> assigned-user
    display name) to resolve hosts that carry no assigned_to. Without it,
    those hosts show up as UNRESOLVED so you know precisely what's missing.
  - an AD_Users export (display name -> email) to resolve names to emails -
    without it, nothing can be resolved to an email at all.
  - an optional Overrides file (name -> email) for the exceptions the fuzzy
    AD match can't get; always wins when present.

Outputs (in output/):
  - noncompliant_consolidated.xlsx   (Worklist + BU Summary)
  - notifications_preview.xlsx       (Notifications + Unresolved)  <- nothing sent

Run:
  python consolidate_noncompliant.py            # read data/, build outputs
  python consolidate_noncompliant.py --regen    # rewrite mock data first
"""

from __future__ import annotations

import csv
import html
import io
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, date

try:
    import openpyxl
    HAVE_XLSX = True
except ImportError:
    HAVE_XLSX = False


# ===========================================================================
# CONFIG
# ===========================================================================

DATA_DIR = os.environ.get("COMPLIANCE_DATA_DIR", "data")
OUTPUT_DIR = os.environ.get("COMPLIANCE_OUTPUT_DIR", "output")

# These reference versions go stale the moment CrowdStrike/Purview ship a new
# client - there's no API call here to auto-refresh them. THRESHOLDS_VERIFIED
# is a manual "confirmed correct as of this date" marker; main() warns loudly
# once it's more than THRESHOLDS_STALE_AFTER_DAYS old (see
# _warn_if_thresholds_stale), so staleness gets noticed instead of silently
# mis-flagging (or failing to flag) real devices indefinitely. Bump both the
# values below AND this date whenever they're re-checked against the actual
# current releases.
THRESHOLDS_VERIFIED = date(2026, 7, 9)
THRESHOLDS_STALE_AFTER_DAYS = 90

CS_LATEST = {"windows": "7.35.20709", "mac": "7.35.20704", "linux": "7.35.18803"}
PURVIEW_LATEST = {
    "windows": {"mocamp": "4.18.26040.7", "engine": "1.1.26020.3"},
    "macos": {"mocamp": "101.26042.0020", "engine": "not tracked"},
}

# --- notification settings -------------------------------------------------
FROM_TEAM = "IT Compliance"     # appears in the message signature

# What the END USER is told to do. Deliberately not the per-finding technical
# 'action' text (sensor versions, Software Center steps, DLP enrollment,
# ZIA/MOCAMP jargon, CMDB inventory upkeep) - that's for whoever runs this
# script, and stays fully intact in the Worklist (noncompliant_consolidated.xlsx)
# 'Action'/'Detail' columns. Fixing this needs hands-on IT servicing, so the
# instruction is to bring the device in - not a self-service fix.
USER_FACING_ACTION = ("Please bring this device to Tech Hub (11/F, AIA Building) "
                      "as soon as possible so it can be serviced and brought back "
                      "into compliance.")

# Host -> assigned-user NAME, from the CMDB export (drop into data/, 'cmdb' in
# the filename). Join key is 'Name' (hostname); 'Assigned to' is a DISPLAY NAME,
# not an email - it gets resolved to an address via AD_Users below.
CMDB_MAPPING = {
    "stem": "cmdb",
    "map": {"hostname": "Name", "name": "Assigned to"},
}

# Directory export that turns a name into an email. Drop into data/ with
# 'ad_users' in the filename. CMDB 'Assigned to' and AD 'DisplayName' use
# DIFFERENT conventions (e.g. "Chan, Tai Man Terry" vs "Chan, Terry-TM"), so the
# match is fuzzy - see resolve_name_to_email(). Only confident matches are used
# to send; ambiguous ones go to a review list, never a guessed email.
# If the sheet has separate GivenName/Surname columns, those are AUTHORITATIVE
# and used directly (no guessing at how DisplayName is formatted). If a row
# lacks them, we fall back to parsing DisplayName for just that row.
AD_USERS = {
    "stem": "ad_users",
    "map": {"name": "DisplayName", "email": "EmailAddress",
            "given": "GivenName", "surname": "Surname"},
}

# Manual name -> email overrides for the exceptions the fuzzy match can't get
# (non-standard names, externals, etc). Optional file, 'overrides' in the name.
# These are AUTHORITATIVE: an override always wins. This is how you fix a
# mis/less-resolved name once and have it stick.
OVERRIDES = {
    "stem": "overrides",
    "map": {"name": "Name", "email": "Email"},
}

# Only these confidence levels are turned into an actual email. Anything lower
# is held for human review rather than risking the wrong recipient.
NOTIFY_CONFIDENCE = {"high", "medium"}


# ===========================================================================
# FILE REGISTRY
# One entry per known report. `meta` records what the file IS; `map` translates
# that file's column names -> our canonical field names. New report -> new
# entry, nothing else changes. Matching is by filename stem, case-insensitive.
# ===========================================================================

# "columns": the exact real header order for that report. Used only to
# self-heal a corrupted header cell (a real export had column 8's name
# replaced with the literal number 0) - see _heal_headers(). Also doubles as
# the header row generate_mock_data() writes, so mock and prod can't drift.
FILE_REGISTRY = {
    # Real file: '20260715_AIAGO-17. Workstation Security Agent Deployment-
    # Crowdstrike.csv'. key "crowdstrike" (not "aiago_workstation_cs")
    # deliberately - same reasoning as "dlp"/"zapp"/"encryption" below:
    # find_dataset() matches by substring against the real filename, which
    # contains "crowdstrike" but never "aiago_workstation_cs".
    #
    # This export lists EVERY device (Windows workstations only - confirmed:
    # os/sys_class_name/chassis_type show zero Mac or Server rows), not just
    # non-compliant ones, so is_compliant_text() gates it here too, on its
    # own 'compliant' column. That column only reflects install presence
    # (crowdstrike_installed), not version currency - a real, small gap
    # (2 of 1519 real rows run a clearly outdated agent_version but are
    # still marked compliant) that was confirmed and deliberately accepted
    # rather than "fixed" by comparing agent_version against CS_LATEST:
    # 468 other rows already run a NEWER build (7.36.20805.0) than CS_LATEST
    # (7.35.20709), meaning CS_LATEST itself is stale - comparing against it
    # directly would falsely flag those 468 as outdated instead.
    #
    # There is no separate Server-CrowdStrike export under the current
    # (dated-CSV) naming convention - the old aiago_server_cs entry (and
    # aiago_windows_purview/aiago_mac_purview, see the "deployment-dlp" entry
    # below) relied on differently-named/schemaed files that no longer exist
    # and have been removed, not merely renamed. Mac DOES have a separate,
    # real export - see "(mac)-cs" below - it just lives under a different
    # ticket (AIAGO-18, not AIAGO-17) with a different schema entirely, so it
    # gets its own registry entry rather than sharing this one.
    "crowdstrike": {
        "meta": {"source": "CrowdStrike", "platform": "Windows", "kind": "Workstation"},
        "map": {
            "bu": "business_unit_code", "hostname": "host_name", "install_status": "install_status",
            "os": "os", "assigned_to": "assigned_to", "last_seen": "last_seen",
            "agent_version": "agent_version", "agent_installed": "crowdstrike",
            "compliance": "compliant", "report_date": "report_date",
        },
        "columns": ["hostname", "business_unit_code", "manufacturer", "chassis_type", "model_id",
                    "serial_number", "company", "assigned_to", "hardware_status", "install_status",
                    "os", "os_domain", "u_vlan", "u_dr_availability", "u_dr_grouping",
                    "u_security_zone", "sys_class_name", "last_discovered", "business_unit",
                    "virtual", "u_non_discoverable_ci", "crowdstrike_required", "crowdstrike",
                    "agent_version", "report_date", "crowdstrike_installed", "host_name",
                    "compliant", "ageing_30_days", "ageing_60_days", "ageing_90_days", "last_seen",
                    "run_at"],
    },
    # Real file: '...Workstation Security Agent Deployment (MAC)-CS.csv'
    # (ticket AIAGO-18, not AIAGO-17 - a separate ticket/export family from
    # the Windows one above). Different schema entirely: no 'crowdstrike'
    # Yes/No column, just a single 'compliant_status' (0/1) flag and no
    # separate hostname duplicate column. No install_status/assigned_to
    # columns either - recipient resolution for this source relies entirely
    # on CMDB hostname->name fallback. cs_issue()'s install-check was
    # widened to also accept "0" (not just "no") specifically to cover this
    # convention - see cs_issue(). Real snapshot: 3/44 non-compliant, all
    # correctly classified as "agent not installed" under that check.
    "(mac)-cs": {
        "meta": {"source": "CrowdStrike", "platform": "Mac", "kind": "Workstation"},
        "map": {
            "bu": "business_unit_code", "hostname": "hostname", "os": "operating_system",
            "last_seen": "cs_last_seen", "agent_version": "cs_agent_version",
            "agent_installed": "compliant_status",
            "compliance": "compliant_status", "report_date": "report_date",
        },
        "columns": ["hostname", "business_unit_code", "agent_type", "compliant_status",
                    "operating_system", "ageing_status", "cs_agent_version", "cs_last_seen",
                    "cs_last_login_user", "serial_number", "model", "operating_system_version",
                    "report_date"],
    },
    # Fuller CMDB-joined DLP/Purview export (same filename family as
    # Crowdstrike above and Zapp below). Real file: 1514 rows, EVERY Windows
    # workstation (both compliant and not) - confirmed comprehensive, not
    # just a supplement - so this is now the SOLE Windows Purview-sourced
    # compliance check. It used to sit alongside separate, thinner
    # aiago_windows_purview/aiago_mac_purview exports and only ADD the hosts
    # those missed (deduped by load_all() on (hostname, source), keeping
    # whichever copy loaded first); those two entries relied on files that no
    # longer exist under the current (dated-CSV) naming convention and have
    # been removed, not merely renamed - this one's real comprehensiveness is
    # exactly why removing them loses no coverage.
    # This export lists EVERY device, not just non-compliant ones, so
    # is_compliant_text() in normalize_file() gates on it before it becomes a finding.
    #
    # key "deployment-dlp" (not bare "dlp"): a separate, real Mac DLP/Purview
    # export exists too (see "(mac)-dlp" below), under the SAME ticket family
    # ('...Workstation Security Agent Deployment (MAC)-DLP.csv') - bare "dlp"
    # would match both real files (confirmed: 'dlp' is a substring of both),
    # and picking the wrong one silently would read a completely incompatible
    # schema. "deployment-dlp" is a substring of the Windows filename
    # ('...Deployment-DLP.csv') but NOT the Mac one ('...Deployment (MAC)-
    # DLP.csv' - the ' (MAC)' in between breaks the contiguous match).
    "deployment-dlp": {
        "meta": {"source": "Purview", "platform": None, "kind": "Workstation"},
        "map": {
            "bu": "business_unit_code", "hostname": "name", "install_status": "install_status",
            "os": "os", "sys_class": "sys_class_name", "assigned_to": "assigned_to",
            "last_seen": "purview_last_seen",
            "mocamp_version": "purview_defender_mocamp_version",
            "engine_version": "purview_defender_engine_version",
            "config_status": "purview_configuration_status",
            "policy_status": "purview_policy_status",
            "compliance": "compliance", "report_date": "report_date",
        },
        "columns": ["name", "manufacturer", "chassis_type", "model_id", "serial_number", "company",
                    "assigned_to", "hardware_status", "install_status", "os", "os_domain", "u_vlan",
                    "u_dr_availability", "u_dr_grouping", "u_security_zone", "sys_class_name",
                    "last_discovered", "virtual", "u_non_discoverable_ci", "u_gis_exclusion",
                    "report_date", "purview_device_name", "purview_configuration_status",
                    "purview_policy_status", "purview_valid_user", "purview_last_seen", "purview_os",
                    "purview_os_version", "purview_last_ip_address", "perview_device_id",
                    "purview_last_policy_sync_time", "purview_is_dlp_enabled",
                    "purview_defender_engine_version", "purview_defender_mocamp_version",
                    "purview_has_dlp_ac_bandwidth_exceeded", "purview_first_time_onboarded",
                    "purview_required", "compliance", "business_unit_code", "ageing_status"],
    },
    # Real file: '...Workstation Security Agent Deployment (MAC)-DLP.csv'
    # (ticket AIAGO-18, same family as "(mac)-cs"/"(mac)-zapp" below).
    # Simpler schema than the Windows DLP export above (no sys_class_name/
    # assigned_to columns at all - hostname is 'intune_computer_name', and
    # recipient resolution for this source relies entirely on CMDB fallback,
    # same as "(mac)-cs" below). Real snapshot: 42/42 rows Compliant (0
    # non-compliant) - genuinely fully compliant today, not a sign of a
    # matching problem; purview_issue() is reused as-is since the columns
    # are the same shape already proven for Windows DLP.
    "(mac)-dlp": {
        "meta": {"source": "Purview", "platform": "Mac", "kind": "Workstation"},
        "map": {
            "bu": "business_unit_code", "hostname": "intune_computer_name",
            "last_seen": "purview_last_seen",
            "mocamp_version": "purview_defender_mocamp_version",
            "engine_version": "purview_defender_engine_version",
            "config_status": "purview_configuration_status",
            "policy_status": "purview_policy_status",
            "compliance": "compliance", "report_date": "report_date",
        },
        "columns": ["intune_computer_name", "intune_orig_bu", "business_unit_code",
                    "intune_file_update", "report_date", "purview_device_name",
                    "purview_configuration_status", "purview_policy_status", "purview_valid_user",
                    "purview_last_seen", "purview_os", "purview_os_version",
                    "purview_last_ip_address", "perview_device_id", "purview_last_policy_sync_time",
                    "purview_is_dlp_enabled", "purview_defender_engine_version",
                    "purview_defender_mocamp_version", "purview_has_dlp_ac_bandwidth_exceeded",
                    "purview_first_time_onboarded", "purview_required", "compliance",
                    "ageing_status"],
    },
    # Zscaler App (client connector) deployment - NOT covered by any of the
    # other reports. This export lists EVERY device (compliant and not)
    # rather than being pre-filtered, so is_compliant_text() gates it too.
    # key "deployment-zapp" (not bare "zapp"): same reasoning as
    # "deployment-dlp" above - a real Mac Zapp export exists too (see
    # "(mac)-zapp" below) and bare "zapp" would match both real filenames.
    "deployment-zapp": {
        "meta": {"source": "Zapp", "platform": None, "kind": "Workstation"},
        "map": {
            "bu": "business_unit_code", "hostname": "hostname", "install_status": "install_status",
            "os": "os", "sys_class": "sys_class_name", "assigned_to": "assigned_to",
            "last_seen": "last_seen_connected_to_zia",
            "zapp_installed": "zapp_installed", "zapp_missing": "zapp_missing",
            "zapp_version": "zapp_version",
            "compliance": "compliant", "report_date": "report_date",
        },
        "columns": ["hostname", "business_unit_code", "zapp", "zapp_required", "zapp_whitelist_bu",
                    "manufacturer", "chassis_type", "model_id", "serial_number", "company",
                    "assigned_to_company", "assigned_to", "install_status", "os", "sys_class_name",
                    "last_discovered", "virtual", "zapp_version", "zapp_user",
                    "last_seen_connected_to_zia", "registration_timestamp", "report_date",
                    "policy_name", "device_state", "last_seen_with_client_connector_active",
                    "zapp_installed", "zapp_missing", "host_name", "compliant", "ageing_30_days",
                    "ageing_60_days", "ageing_90_days", "run_at"],
    },
    # Real file: '...Workstation Security Agent Deployment (MAC)-Zapp.csv'
    # (ticket AIAGO-18). Simpler schema than the Windows Zapp export above -
    # no dedicated zapp_installed/zapp_missing columns, just a single
    # 'compliant_status' (0/1) flag; mapping it directly to zapp_installed
    # works unchanged in zapp_issue(), since "0" is already one of the
    # values that function's install-check accepts (see zapp_issue()).
    # Real snapshot: all 6 non-compliant rows have zero telemetry (blank
    # zapp_version/user/last_seen) - the clean "not installed" shape.
    "(mac)-zapp": {
        "meta": {"source": "Zapp", "platform": "Mac", "kind": "Workstation"},
        "map": {
            "bu": "business_unit_code", "hostname": "hostname", "os": "operating_system",
            "last_seen": "last_seen_connected_to_zia", "zapp_version": "zapp_version",
            "zapp_installed": "compliant_status",
            "compliance": "compliant_status", "report_date": "report_date",
        },
        "columns": ["hostname", "business_unit_code", "agent_type", "compliant_status",
                    "operating_system", "ageing_status", "zapp_version", "zapp_user",
                    "last_seen_connected_to_zia", "serial_number", "model",
                    "operating_system_version", "report_date"],
    },
    # BitLocker (Windows disk encryption) compliance export - NOT covered by
    # any of the other reports. Every row here is a Windows notebook (os,
    # sys_class_name, chassis_type are constant across the real export), so
    # platform is fixed rather than derived. This export lists EVERY device
    # (compliant and not) rather than being pre-filtered, so is_compliant_text()
    # gates it too, on its own 'compliant' column.
    #
    # key "encryption" (not "bitlocker") deliberately: find_dataset() matches
    # by substring against the real filename, e.g.
    # '20260715-AIAGO-19. Hard Disk Encryption Compliance.csv' - which
    # contains "encryption" but never "bitlocker" anywhere. See how "dlp" and
    # "zapp" above are each keyed on a word actually present in their own
    # real filenames, for the same reason.
    "encryption": {
        "meta": {"source": "BitLocker", "platform": "Windows", "kind": "Workstation"},
        "map": {
            "bu": "business_unit_code", "hostname": "host_name", "install_status": "install_status",
            "os": "os", "assigned_to": "assigned_to", "last_seen": "last_discovered",
            "encryption_status": "encryption_status", "setting_state_summary": "setting_state_summary",
            "compliance": "compliant", "report_date": "report_date",
        },
        "columns": ["name_x", "host_name", "manufacturer", "chassis_type", "model_id", "serial_number",
                    "company", "assigned_to", "hardware_status", "install_status", "os", "os_domain",
                    "u_vlan", "u_dr_availability", "u_dr_grouping", "u_security_zone", "sys_class_name",
                    "last_discovered", "virtual", "u_non_discoverable_ci", "business_unit_code",
                    "domain_name", "compliance_status", "exemption", "compliance_status_details",
                    "collection_id", "site_code", "encryption_status", "setting_state_summary",
                    "advanced_bitlocker_state", "policy_details", "compliant", "report_date",
                    "intune_compliance_status", "mbam_compliance_status", "cmdb_bu", "gis_bu",
                    "count_total_encrypted", "total_encrypted", "total_unencrypted", "ageing_30_days",
                    "ageing_60_days", "ageing_90_days", "run_at"],
    },
}

CANON_COLUMNS = [
    "bu", "hostname", "source", "platform", "kind",
    "issue", "action", "detail",
    "last_seen", "assigned_to", "compliance", "report_date", "source_file",
]


# ===========================================================================
# CELL HELPERS + FILE I/O
# ===========================================================================

def cell_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


# ===========================================================================
# DATA SOURCE  (file I/O seam)
# Every INPUT read in this file goes through this abstraction instead of
# calling open()/os.listdir() directly, so swapping local disk for a future
# backend is a contained edit to _get_data_source() below, not a hunt
# through every reader. Deliberately NOT part of this seam: output writing
# (write_worklist/write_notifications_preview/write_html_preview) and mock-
# fixture generation (generate_mock_data/_write_mock) - both stay direct
# disk I/O for now, same as before this refactor.
# ===========================================================================

class DataSource:
    """Abstraction over 'a directory of input report/config files'.
    LocalDataSource and SharePointDataSource (below) are the two
    implementations today - see _get_data_source() for which one actually
    backs a run (currently always LocalDataSource - SharePointDataSource
    exists and is tested, but isn't wired in as active anywhere yet)."""

    def exists(self) -> bool:
        """-> True if this source is reachable at all (e.g. the local
        directory exists / the blob container exists)."""
        raise NotImplementedError

    def list_files(self) -> list:
        """-> sorted list of file names available. These are logical names
        (no directory component), not filesystem paths."""
        raise NotImplementedError

    def read_file(self, name: str) -> bytes:
        """-> the raw bytes of the named file."""
        raise NotImplementedError


class LocalDataSource(DataSource):
    """Reads a directory on local disk - exactly what this file did before
    this abstraction existed (os.listdir() + open().read()), byte for byte.
    Extension point: a future BlobDataSource(container_client) would
    implement exists()/list_files()/read_file() against Azure Blob Storage
    instead - see DataSource above."""

    def __init__(self, directory: str):
        self.directory = directory

    def exists(self) -> bool:
        return os.path.isdir(self.directory)

    def list_files(self) -> list:
        if not self.exists():
            return []
        return sorted(os.listdir(self.directory))

    def read_file(self, name: str) -> bytes:
        with open(os.path.join(self.directory, name), "rb") as f:
            return f.read()


# ===========================================================================
# SHAREPOINT DATA SOURCE  (mock-tested only - see class docstring)
# Small, mockable Graph HTTP helpers, same shape as send_email.py's
# _get_graph_token()/_graph_send_mail() (raw urllib, no new dependency).
# Duplicated in miniature rather than imported from send_email.py:
# send_email.py already imports consolidate_noncompliant.py, so the reverse
# import would be circular, and this file's token helper deliberately takes
# `scope` as a parameter (see get_graph_token_for_scope()) where
# send_email.py's hardcodes Mail.Send's '.default' - a SharePoint-reading
# app registration is not necessarily the same one, or the same granted
# permission.
# ===========================================================================

GRAPH_TOKEN_URL = "https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
GRAPH_BASE_URL = "https://graph.microsoft.com/v1.0"


def get_graph_token_for_scope(tenant_id: str, client_id: str, client_secret: str,
                               scope: str = "https://graph.microsoft.com/.default") -> str:
    """Client-credentials Graph token fetch. `scope` is a parameter, not a
    given: don't assume send_email.py's Mail.Send-scoped token (or that
    app registration) also covers Sites.Selected/Files.Read.All - whatever
    app registration/permission actually ends up granted for file reads may
    be entirely separate."""
    data = urllib.parse.urlencode({
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": scope,
        "grant_type": "client_credentials",
    }).encode()
    req = urllib.request.Request(GRAPH_TOKEN_URL.format(tenant=tenant_id), data=data, method="POST")
    with urllib.request.urlopen(req, timeout=30) as resp:
        payload = json.loads(resp.read().decode())
    return payload["access_token"]


def _graph_get_bytes(url: str, token: str) -> bytes:
    """Raw authenticated GET -> response bytes. The one HTTP primitive both
    _graph_list_children() and _graph_download_content() build on, so both
    are easily mockable in tests without ever making a real call."""
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read()


def _graph_list_children(drive_id: str, path: str, token: str) -> list:
    """-> list of Graph driveItem dicts (each has at least 'name' and either
    'folder' or 'file') for the children of `path` within `drive_id`. Single
    page only - real folder/file counts here (a handful of dated subfolders,
    a handful of report files per day) are small enough that this hasn't
    needed @odata.nextLink pagination handling; flagged here rather than
    silently assumed away, since there's no live data yet to confirm that."""
    url = (f"{GRAPH_BASE_URL}/drives/{drive_id}/root:/"
           f"{urllib.parse.quote(path, safe='/')}:/children")
    payload = json.loads(_graph_get_bytes(url, token).decode())
    return payload.get("value", [])


def _graph_download_content(drive_id: str, path: str, token: str) -> bytes:
    url = (f"{GRAPH_BASE_URL}/drives/{drive_id}/root:/"
           f"{urllib.parse.quote(path, safe='/')}:/content")
    return _graph_get_bytes(url, token)


class SharePointDataSource(DataSource):
    """Reads report files from a SharePoint document library via Microsoft
    Graph, resolving the latest dated (YYYY/MM/DD) subfolder under a base
    path DYNAMICALLY rather than assuming any particular date or cadence -
    new folders appear irregularly (usually weekly, no guaranteed day of the
    week), so hardcoding today's date or a fixed weekday offset would
    silently miss folders, or worse, silently read a stale one. See
    _resolve_latest_dated_folder().

    NOT wired in anywhere as the active DataSource - _get_data_source()
    below still only ever returns LocalDataSource. There is no live
    Sites.Selected/Files.Read.All grant yet, so this class is mock-tested
    only (every test replaces _graph_list_children/_graph_download_content,
    never makes a real HTTP call) - built and ready to activate the moment
    that grant lands, same shape as send_email.py's DatabricksSecretSource.

    drive_id: the target document library's Graph drive ID. Not resolved
    from a site hostname/path here, since which site this actually lives on
    isn't confirmed yet either - supply it directly once it is.
    base_path: the folder ABOVE the YYYY/MM/DD structure, e.g.
    'Documents/Reports-Prod/AIAGO/Weekly Dashboard/17. Workstation Security
    Agent Deployment' - taken exactly as given, no assumption about whether
    a leading 'Documents/' needs stripping.
    get_token: zero-arg callable -> a valid Graph access token. An opaque
    injection point deliberately, not a call to get_graph_token_for_scope()
    itself - so this class never assumes which tenant/app-registration/scope
    actually ends up with the read permission; the caller decides.
    """

    def __init__(self, drive_id: str, base_path: str, get_token):
        self.drive_id = drive_id
        self.base_path = base_path.strip("/")
        self._get_token = get_token
        self._folder = None

    def _latest_numeric_subfolder(self, path: str) -> str:
        """-> the name of `path`'s numeric-named FOLDER child with the
        highest INTEGER value - e.g. among ['9', '10'] -> '10', not '9',
        which a plain string/alphabetical sort would wrongly prefer (the
        exact unpadded-name edge case this must not get wrong). Raises if no
        numeric-named folder exists, so a genuinely unexpected structure
        surfaces as a clear error instead of silently resolving to nothing."""
        children = _graph_list_children(self.drive_id, path, self._get_token())
        numeric = [c["name"] for c in children
                   if c.get("folder") is not None and c.get("name", "").isdigit()]
        if not numeric:
            raise RuntimeError(
                f"SharePoint: no numeric (dated) subfolder found under {path!r} - "
                f"found: {[c.get('name') for c in children] or '(empty)'}")
        return max(numeric, key=int)

    def _resolve_latest_dated_folder(self) -> str:
        """-> full path to the latest YYYY/MM/DD folder under self.base_path.
        Three separate 'list children, pick the max' calls - year, then
        month within that year, then day within that month - never a guess,
        never today's date, never an assumed cadence."""
        year = self._latest_numeric_subfolder(self.base_path)
        year_path = f"{self.base_path}/{year}"
        month = self._latest_numeric_subfolder(year_path)
        month_path = f"{year_path}/{month}"
        day = self._latest_numeric_subfolder(month_path)
        return f"{month_path}/{day}"

    def _resolved_folder(self) -> str:
        """Resolved lazily on first use (not in __init__) and cached for
        this instance's lifetime - matches DatabricksSecretSource's lazy,
        per-instance client caching in send_email.py. _get_data_source()
        already constructs a fresh DataSource per call, so a fresh
        SharePointDataSource re-resolves the latest folder next time too."""
        if self._folder is None:
            self._folder = self._resolve_latest_dated_folder()
        return self._folder

    def exists(self) -> bool:
        try:
            self._resolved_folder()
            return True
        except Exception:
            return False

    def list_files(self) -> list:
        folder = self._resolved_folder()
        children = _graph_list_children(self.drive_id, folder, self._get_token())
        return sorted(c["name"] for c in children if c.get("file") is not None)

    def read_file(self, name: str) -> bytes:
        folder = self._resolved_folder()
        return _graph_download_content(self.drive_id, f"{folder}/{name}", self._get_token())


def _get_data_source() -> DataSource:
    """The single place that decides which DataSource backs every input
    read. Constructed fresh on every call (not cached at import time) so it
    always reflects the CURRENT value of DATA_DIR - important because tests
    (and DATA_DIR itself) can change at runtime. Always LocalDataSource today;
    switching to SharePointDataSource (already built and tested - see its
    class docstring) once its Graph permission grant lands is a one-line
    change here, not a hunt through every reader."""
    return LocalDataSource(DATA_DIR)


def _read_csv_rows(name: str) -> list:
    raw = _get_data_source().read_file(name)
    try:
        text = raw.decode("utf-8-sig")
    except Exception as e:
        _reraise_with_filename(name, e)
    return list(csv.DictReader(io.StringIO(text)))


def _looks_like_corruption(value: str) -> bool:
    """True only for values that couldn't plausibly be a deliberate column
    name - blank, or purely numeric (the actual corruption seen in
    production: a header cell overwritten with the literal number 0). A
    value containing any letters (e.g. a genuine rename to 'foo_v2') is left
    alone - see _heal_headers()."""
    v = (value or "").strip()
    if not v:
        return True
    try:
        float(v)
        return True
    except ValueError:
        return False


def _heal_headers(headers: list, expected: list, where: str = "") -> list:
    """Recover a header that got clobbered at the source (seen in the wild:
    'proc_cs_version_status' exported as the literal number 0). Only heals a
    slot when: the actual value doesn't look like a plausible column name at
    all (blank or purely numeric - see _looks_like_corruption; a genuine
    rename to a real-looking new name is deliberately left alone instead of
    being silently rewritten back, which would hide that anything changed),
    the actual name doesn't match ANY expected column (so a rename that's
    really just two columns swapped is never touched), and the expected name
    is otherwise missing entirely from the row (so a genuine reorder is never
    misaligned)."""
    if not expected or len(headers) != len(expected):
        return headers
    healed = list(headers)
    changed = False
    for i, exp in enumerate(expected):
        if (headers[i] != exp and _looks_like_corruption(headers[i])
                and headers[i] not in expected and exp not in headers):
            healed[i] = exp
            changed = True
            print(f"  ! header self-heal {where}: column {i} read as {headers[i]!r}; "
                  f"expected file layout says {exp!r} - using that")
    return healed if changed else headers


def _require_openpyxl(path: str) -> None:
    """openpyxl is optional at import time (HAVE_XLSX) so an all-CSV data/
    directory works without it, but the two call sites that actually touch
    .xlsx files skipped this check, so a missing openpyxl surfaced as a bare
    NameError deep in a stack trace the moment an .xlsx file was involved,
    instead of a message saying what to install."""
    if not HAVE_XLSX:
        raise RuntimeError(
            f"openpyxl is required to read {path!r} but isn't installed in this "
            f"Python environment (pip install openpyxl, or add it to requirements.txt "
            f"for this runtime). CSV-only data/ directories work without it."
        )


def _reraise_with_filename(name: str, exc: Exception) -> None:
    """Re-raises `exc` with `name` folded into the message, so a read failure
    deep in openpyxl/csv (e.g. a corrupted .xlsx, or a non-UTF-8 .csv) names
    the file that caused it instead of a bare 'File is not a zip file' with
    nothing to locate it by in a headless run reading several input files.
    Preserves the exact exception type and chains the original via `from exc`
    - this only changes what the message says, never what failed or whether
    it's raised at all. UnicodeDecodeError needs its own branch: unlike most
    exceptions, its constructor takes (encoding, object, start, end, reason)
    rather than a plain message string, and its str() is generated from those
    fields - a filename can only be folded in via `reason`."""
    if isinstance(exc, UnicodeDecodeError):
        raise UnicodeDecodeError(
            exc.encoding, exc.object, exc.start, exc.end, f"{name}: {exc.reason}"
        ) from exc
    raise type(exc)(f"{name}: {exc}") from exc


def _open_xlsx_headers(name: str, sheet: str, expected_headers: list):
    """Opens the workbook, selects the sheet, and reads+heals just the
    header row. -> (workbook, row_iterator, headers). Caller must close
    `workbook`. `row_iterator` continues from the row AFTER headers, for a
    caller that wants the data rows too (_read_xlsx_rows) - a caller that
    only wants headers (_read_headers) can close and return right away
    without ever touching `row_iterator`, so pulling this logic out into a
    shared helper does NOT cost reading the rest of the sheet."""
    data = _get_data_source().read_file(name)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        _reraise_with_filename(name, e)
    ws = wb[sheet] if sheet else wb.active
    it = ws.iter_rows(values_only=True)
    try:
        headers = [cell_to_str(h) for h in next(it)]
    except StopIteration:
        return wb, it, []
    headers = _heal_headers(headers, expected_headers, where=os.path.basename(name))
    return wb, it, headers


def _read_xlsx_rows(name: str, sheet: str = None, expected_headers: list = None) -> list:
    _require_openpyxl(name)
    wb, it, headers = _open_xlsx_headers(name, sheet, expected_headers)
    if not headers:
        wb.close()
        return []
    out = []
    for row in it:
        if row is None or all(c is None for c in row):
            continue
        out.append({h: (row[i] if i < len(row) else None) for i, h in enumerate(headers)})
    wb.close()
    return out


def _read_any(name: str, sheet: str = None, expected_headers: list = None) -> list:
    if name.lower().endswith(".xlsx"):
        return _read_xlsx_rows(name, sheet, expected_headers)
    return _read_csv_rows(name)


def _read_headers(name: str, sheet: str = None, expected_headers: list = None) -> list:
    """Just the header row, for the pre-flight column check. Deliberately
    does not read the rest of the sheet - see _open_xlsx_headers()."""
    if name.lower().endswith(".xlsx"):
        _require_openpyxl(name)
        wb, it, headers = _open_xlsx_headers(name, sheet, expected_headers)
        wb.close()
        return headers
    raw = _get_data_source().read_file(name)
    try:
        text = raw.decode("utf-8-sig")
    except Exception as e:
        _reraise_with_filename(name, e)
    try:
        return [cell_to_str(h) for h in next(csv.reader(io.StringIO(text)))]
    except StopIteration:
        return []


# ===========================================================================
# REASON -> ISSUE + ACTION
# ===========================================================================

def cs_issue(reason: str, agent_installed: str = "") -> tuple:
    inst = (agent_installed or "").strip().lower()
    r = (reason or "").strip().lower()
    # 'agent not installed' is the authoritative signal: some files (Workstation)
    # leave the reason blank for agent-less hosts instead of writing "Unknown".
    # Accepts "0"/"false" too, not just "no": the Mac CrowdStrike export
    # ("(mac)-cs" in FILE_REGISTRY) has no Yes/No install column at all, only
    # a "0"/"1" compliant_status flag mapped straight into agent_installed -
    # same convention zapp_issue() already accepts for the same reason.
    if inst in ("no", "0", "false") or r == "unknown":
        return ("CrowdStrike agent not installed",
                "Install the CrowdStrike sensor (latest from the Prod share)")
    if r == "outdated":
        return ("CrowdStrike agent outdated",
                f"Update sensor to latest (Win {CS_LATEST['windows']} / "
                f"Mac {CS_LATEST['mac']} / Linux {CS_LATEST['linux']})")
    if r == "latest":
        return ("Agent current but NOT reporting",
                "Check network connectivity / power the machine on so it reports")
    if not r:
        return ("CrowdStrike status not reported",
                "Verify agent status on the host; refer to remediation guidance")
    return (f"CrowdStrike status: {reason}", "Refer to remediation guidance")


def purview_issue(config_status: str, policy_status: str, platform: str = "",
                  mocamp: str = "", engine: str = "") -> tuple:
    cfg = (config_status or "").strip()
    pol = (policy_status or "").strip()
    detail = (f"config={cfg or 'n/a'}, policy={pol or 'n/a'}, "
              f"mocamp={mocamp or 'n/a'}, engine={engine or 'n/a'}")
    ref = PURVIEW_LATEST.get("macos" if platform == "Mac" else "windows", {})
    statuses = {cfg.lower(), pol.lower()}

    if "notupdated" in statuses:
        ref_txt = f"mocamp {ref.get('mocamp', '?')} / engine {ref.get('engine', '?')}"
        return ("Microsoft Defender / Purview components not updated",
                f"Update Defender to the reference versions ({ref_txt}); "
                f"confirm 'Purview DLP Enrollment' in Software Center",
                detail)
    if not cfg and not pol:
        # Blank telemetry across the Purview columns = host has no Purview data,
        # i.e. it isn't onboarded / reporting to Purview DLP. This is the
        # majority case and matches the email's Software Center enrollment step.
        return ("Purview DLP not enrolled / not reporting",
                "In Software Center, check 'Purview DLP Enrollment' and click "
                "Install to enroll; follow the Onboarding / Troubleshooting deck",
                detail)
    return (f"Purview status: config={cfg or 'n/a'}, policy={pol or 'n/a'}",
            "Follow the Onboarding / Troubleshooting deck",
            detail)


def is_compliant_text(value: str) -> bool:
    """True if a source's own 'compliance' column says this row is already
    compliant. The original 5 AIAGO exports are pre-filtered to non-compliant
    rows only (this is always False for them), but the fuller unfiltered
    exports (Zapp, the merged DLP export, BitLocker) list EVERY device, so
    normalize_file gates on this before generating a finding - otherwise a
    compliant device would get reported as a finding too."""
    return (value or "").strip().lower() in ("1", "true", "yes", "compliant")


def zapp_issue(zapp_installed: str, zapp_missing: str) -> tuple:
    if (zapp_missing or "").strip().lower() in ("1", "true", "yes") or \
       (zapp_installed or "").strip().lower() in ("0", "false", "no"):
        return ("Zapp (Zscaler Client Connector) not installed",
                "Install Zapp from Software Center; confirm it registers and connects to ZIA")
    return ("Zapp reporting non-compliant",
            "Verify Zapp registration/connectivity; refer to remediation guidance")


def bitlocker_issue(encryption_status: str, setting_state_summary: str) -> tuple:
    """Real data (data/*Hard Disk Encryption Compliance*.csv, 33 non-compliant/
    blank rows out of 1380) splits into three shapes:
      - encryption_status='notEncrypted' (4 rows): the clear case - device
        reports in, drive isn't encrypted.
      - encryption_status='encrypted' but setting_state_summary='notAssigned'
        (1 row): encrypted, but the compliance policy itself isn't applied to
        the device, so its state isn't actually being evaluated.
      - encryption_status blank (28 rows, the majority): no BitLocker/Intune
        telemetry at all for this host, matching CrowdStrike's own
        'status not reported' shape for the same root cause (host not
        reporting in) - see cs_issue().
    """
    enc = (encryption_status or "").strip().lower()
    summ = (setting_state_summary or "").strip().lower()
    if enc == "notencrypted":
        return ("BitLocker drive encryption not enabled",
                "Enable BitLocker drive encryption on this device (Settings > Device "
                "encryption, or via Software Center) and confirm the recovery key "
                "escrows to Intune")
    if not enc:
        return ("BitLocker status not reported",
                "Verify Intune/MBAM reporting is active on this host; refer to "
                "remediation guidance")
    if summ == "notassigned":
        return ("BitLocker encrypted but compliance policy not applied",
                "Confirm the BitLocker compliance policy is assigned to this device in Intune")
    return (f"BitLocker status: encryption={encryption_status or 'n/a'}, "
            f"setting={setting_state_summary or 'n/a'}",
            "Refer to remediation guidance")


# ===========================================================================
# LOAD + NORMALIZE
# ===========================================================================

def _match_registry(stem: str):
    key = stem.lower()
    if key in FILE_REGISTRY:
        return key, FILE_REGISTRY[key]
    for rk, entry in FILE_REGISTRY.items():
        if key.startswith(rk) or rk in key:
            return rk, entry
    return None, None


def _platform_from_os(os_text: str) -> str:
    t = (os_text or "").lower()
    if "win" in t:
        return "Windows"
    if "mac" in t or "osx" in t or "darwin" in t:
        return "Mac"
    if "linux" in t or "rhel" in t or "ubuntu" in t or "centos" in t:
        return "Linux"
    return "Unknown"


def normalize_file(path: str, sheet: str = None, registry_key: str = None) -> list:
    stem = sheet or os.path.splitext(os.path.basename(path))[0]
    if registry_key:
        entry = FILE_REGISTRY[registry_key]
    else:
        _, entry = _match_registry(stem)
    label_src = f"{os.path.basename(path)}" + (f" [{sheet}]" if sheet else "")
    if not entry:
        print(f"  ! skipped (unknown report): {label_src}")
        return []

    cmap, meta = entry["map"], entry["meta"]
    rows = []
    for raw in _read_any(path, sheet, expected_headers=entry.get("columns")):
        f = {canon: cell_to_str(raw.get(col, "")) for canon, col in cmap.items()}
        # The original 5 exports are pre-filtered to non-compliant rows, so
        # this is always False for them. The fuller unfiltered exports (Zapp,
        # the merged DLP export) list EVERY device, so skip the ones their
        # own 'compliance' column already says are fine - a no-op for the
        # pre-filtered sources since 'compliance' there is always non-compliant.
        if is_compliant_text(f.get("compliance")):
            continue

        # Platform: fixed for most files. For servers it comes from ser_os, but
        # that column is sometimes empty while ser_sys_class_name carries the
        # OS signal (e.g. "Citrix VDI", "Linux Server"); try os first, then
        # fall back to sys_class so those rows aren't mislabelled "Unknown".
        platform = meta["platform"]
        if not platform:
            platform = _platform_from_os(f.get("os"))
            if platform == "Unknown":
                platform = _platform_from_os(f.get("sys_class"))
            if platform == "Unknown":
                # neither column buckets to Win/Mac/Linux (e.g. "Citrix VDI");
                # keep the most specific raw label rather than a bare "Unknown".
                platform = f.get("os") or f.get("sys_class") or "Unknown"
        if meta["source"] == "CrowdStrike":
            issue, action = cs_issue(f.get("cs_reason"), f.get("agent_installed"))
            detail = f"reason={f.get('cs_reason') or 'n/a'}, installed={f.get('agent_installed') or 'n/a'}"
            if f.get("agent_version"):
                detail += f", agent={f['agent_version']}"
        elif meta["source"] == "Zapp":
            issue, action = zapp_issue(f.get("zapp_installed"), f.get("zapp_missing"))
            detail = (f"installed={f.get('zapp_installed') or 'n/a'}, "
                      f"missing={f.get('zapp_missing') or 'n/a'}, "
                      f"version={f.get('zapp_version') or 'n/a'}")
        elif meta["source"] == "BitLocker":
            issue, action = bitlocker_issue(f.get("encryption_status"), f.get("setting_state_summary"))
            detail = (f"encryption={f.get('encryption_status') or 'n/a'}, "
                      f"setting={f.get('setting_state_summary') or 'n/a'}")
        else:
            issue, action, detail = purview_issue(
                f.get("config_status"), f.get("policy_status"),
                platform, f.get("mocamp_version"), f.get("engine_version"))

        rows.append({
            "bu": f.get("bu") or "(no BU)",
            "hostname": f.get("hostname"),
            "source": meta["source"],
            "platform": platform,
            "kind": meta["kind"],
            "issue": issue,
            "action": action,
            "detail": detail,
            "last_seen": f.get("last_seen"),
            "assigned_to": f.get("assigned_to", ""),
            "compliance": f.get("compliance"),
            "report_date": f.get("report_date"),
            "source_file": label_src,
        })
    print(f"  + {label_src:40s} {len(rows):4d} rows  [{meta['source']}/{meta['kind']}]")
    return rows


def load_all() -> list:
    """Find each of the known reports, whether it's its own file or a tab
    inside a larger multi-tab workbook (e.g. everything living in one
    'CompliantReport(Working).xlsx'). Each report is loaded exactly once, by
    whichever form is found first (standalone file takes priority).

    Rows are deduplicated by (hostname, source), keeping whichever copy was
    loaded first - defense against two registry entries both covering the
    same report under the same source (this used to matter for real: "dlp"
    was once a fuller re-export overlapping aiago_windows_purview/
    aiago_mac_purview - see the "dlp" FILE_REGISTRY comment for why those
    were removed, not just renamed). No two current entries share a source,
    so this dedup is dormant today, not exercised - kept as-is since it's
    free insurance against the same situation recurring, e.g. if a report
    ever splits back into two overlapping files."""
    all_rows = []
    print(f"Reading reports from '{DATA_DIR}/':")
    for rk in FILE_REGISTRY:
        found = find_dataset(rk)
        if not found:
            print(f"  ! not found (file or tab): {rk}")
            continue
        path, sheet = found
        all_rows.extend(normalize_file(path, sheet, registry_key=rk))

    seen, deduped, dropped = set(), [], 0
    for r in all_rows:
        key = ((r.get("hostname") or "").strip().upper(), r["source"])
        if key in seen:
            dropped += 1
            continue
        seen.add(key)
        deduped.append(r)
    if dropped:
        print(f"  (dropped {dropped} row(s) already covered by an earlier-loaded "
              f"report for the same host+source)")
    return deduped


# ===========================================================================
# CONSOLIDATED WORKLIST OUTPUT
# ===========================================================================

SOURCE_KINDS = ["CrowdStrike", "Purview", "Zapp"]  # summary columns in write_worklist()


def summarize_by_bu(rows: list) -> dict:
    bus = {}
    for r in rows:
        b = bus.setdefault(r["bu"], {"total": 0, "Workstation": 0, "Server": 0,
                                     **{s: 0 for s in SOURCE_KINDS}})
        b["total"] += 1
        b[r["source"]] = b.get(r["source"], 0) + 1
        b[r["kind"]] = b.get(r["kind"], 0) + 1
    return dict(sorted(bus.items()))


def write_worklist(rows: list) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if not HAVE_XLSX:
        path = os.path.join(OUTPUT_DIR, "noncompliant_consolidated.csv")
        with open(path, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=CANON_COLUMNS)
            w.writeheader()
            for r in sorted(rows, key=lambda x: (x["bu"], x["source"], x["hostname"])):
                w.writerow({c: r.get(c, "") for c in CANON_COLUMNS})
        return path

    path = os.path.join(OUTPUT_DIR, "noncompliant_consolidated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worklist"
    ws.append([c.replace("_", " ").title() for c in CANON_COLUMNS])
    for r in sorted(rows, key=lambda x: (x["bu"], x["source"], x["hostname"])):
        ws.append([r.get(c, "") for c in CANON_COLUMNS])
    ws.freeze_panes = "A2"

    summary = wb.create_sheet("BU Summary")
    summary.append(["Business Unit", "Total"] + SOURCE_KINDS + ["Workstation", "Server"])
    for bu, s in summarize_by_bu(rows).items():
        summary.append([bu, s["total"]] + [s[k] for k in SOURCE_KINDS] + [s["Workstation"], s["Server"]])
    summary.freeze_panes = "A2"
    wb.save(path)
    return path


# ===========================================================================
# RECIPIENT RESOLUTION
# ===========================================================================

# ===========================================================================
# RECIPIENT RESOLUTION   hostname -> name (CMDB) -> email (AD_Users, fuzzy)
# ===========================================================================

def _list_sheets(name: str) -> list:
    if not name.lower().endswith(".xlsx"):
        return [None]
    _require_openpyxl(name)
    data = _get_data_source().read_file(name)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    except Exception as e:
        _reraise_with_filename(name, e)
    names = wb.sheetnames
    wb.close()
    return names


# ===========================================================================
# DATASET LOCATOR
# A "dataset" (a report, or CMDB, or AD_Users, or Overrides) can arrive two
# ways: as its OWN file (data/CMDB_Mapping.xlsx), or as ONE SHEET inside a
# larger multi-tab workbook (e.g. "CompliantReport(Working).xlsx" with a
# 'CMDB' tab and an 'AD_Users' tab). find_dataset() checks both, by matching
# the stem keyword against filenames first, then against sheet names inside
# every .xlsx in data/. This is the ONLY place that needs to know which shape
# the input actually took.
# ===========================================================================

_LEADING_DATE_RE = re.compile(r"^(\d{8})")


def _leading_date_key(filename: str):
    """-> the int value of a leading YYYYMMDD date prefix real report
    exports carry (e.g. '20260715_AIAGO-17...', '20260715-AIAGO-19...'), or
    None if the filename doesn't start with 8 digits at all (e.g.
    'AIAGO_Workstation_CS.xlsx', or a test fixture like
    'Alpha_DLP_Extra.xlsx'). Parsed as an int, not compared as a string, so
    this is the actual date value - not, say, '20260715' being treated as
    "less than" some unrelated 9-digit prefix."""
    m = _LEADING_DATE_RE.match(os.path.basename(filename))
    return int(m.group(1)) if m else None


def _pick_latest_dated(matches: list) -> str:
    """Tie-breaker for find_dataset() when more than one file matches the
    same keyword. A real, now-observed case: successive weekly exports of
    the same report (e.g. DLP, Zapp) can sit in data/ at once - a 20260703
    copy alongside a 20260715 one - since old exports aren't necessarily
    cleaned out before the next one lands. Prefers the candidate with the
    LATEST such date, compared as an integer - the same "parse and compare
    the actual number, don't string-sort it" principle SharePointDataSource's
    folder discovery uses, for the same reason (e.g. a 202607 prefix must
    not lose to some unrelated string that happens to sort later). Falls
    back to `matches[0]` (already alphabetically-first, since callers build
    `matches` from a sorted listing) when none of the candidates carry a
    recognizable date prefix, leaving that already-tested, deterministic
    behavior unchanged for undated ambiguous matches."""
    dated = [(m, _leading_date_key(m)) for m in matches]
    if not any(key is not None for _, key in dated):
        return matches[0]
    return max(dated, key=lambda pair: (pair[1] is not None, pair[1] or 0))[0]


def find_dataset(stem_keyword: str):
    """-> (filename, sheet_name_or_None) for the match, else None. `filename`
    is a logical name within the active DataSource (see _get_data_source()),
    not necessarily a filesystem path.

    Matching is by substring against filenames first, then sheet tabs inside
    every .xlsx - necessarily loose, since real report filenames carry
    timestamp/ticket prefixes this needs to see through. That looseness has a
    real cost: it can false-positive on an unrelated file/sheet that happens
    to contain the same keyword (e.g. 'dlp' matching a stray meeting-notes
    file). Several things keep that from being a silent, unreproducible
    surprise:
      - candidates are scanned in SORTED order, not whatever order the
        DataSource happens to return (which, for LocalDataSource, mirrors
        the filesystem/OS-dependent, not-guaranteed-stable os.listdir()
        order), so the same input directory always resolves the same way.
      - if MORE THAN ONE file matches the same keyword, that's printed as a
        warning naming every candidate. The one actually used is the
        latest-dated candidate if any of them carry a leading YYYYMMDD
        prefix - see _pick_latest_dated() - compared as an integer, never a
        string; otherwise (no candidate has a recognizable date prefix) it
        falls back to the alphabetically-first one, same as before. Sheet-
        tab matches (inside a multi-tab workbook) keep the plain
        alphabetically-first rule unconditionally - there's no evidence
        sheet-tab names use this dated-export naming convention at all.
    This does NOT solve the other real risk - a renamed report simply
    matching nothing and silently vanishing - which is why validate_headers()
    surfaces every "no matching file or sheet found" into main()'s end-of-run
    banner rather than letting it stay a single easy-to-miss console line.
    """
    source = _get_data_source()
    if not source.exists():
        return None
    key = stem_keyword.lower()
    xlsx_files = []
    filename_matches = []
    for fn in source.list_files():
        if not fn.lower().endswith((".xlsx", ".csv")):
            continue
        base = os.path.splitext(fn)[0].lower()
        # 1. filename itself matches -> standalone file, own active sheet
        if base.startswith(key) or key in base:
            filename_matches.append(fn)
        if fn.lower().endswith(".xlsx"):
            xlsx_files.append(fn)

    if filename_matches:
        if len(filename_matches) > 1:
            chosen = _pick_latest_dated(filename_matches)
            others = ", ".join(f for f in filename_matches if f != chosen)
            why = "latest dated" if _leading_date_key(chosen) is not None else "alphabetically first"
            print(f"  ! '{stem_keyword}' matches {len(filename_matches)} files in "
                  f"'{DATA_DIR}/' - using {chosen} ({why}), ignoring: {others}")
        else:
            chosen = filename_matches[0]
        return chosen, None

    # 2. no filename matched -> look for a matching TAB inside every workbook.
    # Sheet tabs likely won't carry the "aiago_" file-prefix, so compare with
    # that stripped, and match in either direction (key-in-sheet or
    # sheet-in-key) since tab names are often abbreviated.
    core_key = key.replace("_", "").replace("aiago", "")
    sheet_matches = []
    for fn in xlsx_files:
        for sheet in _list_sheets(fn):
            s = (sheet or "").lower().replace(" ", "").replace("_", "")
            if len(s) >= 3 and (core_key in s or s in core_key):
                sheet_matches.append((fn, sheet))
    if sheet_matches:
        if len(sheet_matches) > 1:
            others = ", ".join(f"{fn} [{s}]" for fn, s in sheet_matches[1:])
            first_fn, first_sheet = sheet_matches[0]
            print(f"  ! '{stem_keyword}' matches {len(sheet_matches)} sheets - using "
                  f"{first_fn} [{first_sheet}], ignoring: {others}")
        return sheet_matches[0]
    return None


def read_cmdb_mapping() -> dict:
    """hostname (UPPER) -> assigned-user display name (raw)."""
    found = find_dataset(CMDB_MAPPING["stem"])
    if not found:
        print("No CMDB file/sheet in data/ - CS-only workstations will be unresolved.")
        return {}
    path, sheet = found
    m = CMDB_MAPPING["map"]
    out = {}
    for raw in _read_any(path, sheet):
        host = cell_to_str(raw.get(m["hostname"], "")).upper()
        name = cell_to_str(raw.get(m["name"], ""))
        if host and name:
            out[host] = name
    where = f"{os.path.basename(path)}" + (f" [{sheet}]" if sheet else "")
    print(f"CMDB '{where}': {len(out)} host->name entries.")
    return out


# --- name normalization / parsing ------------------------------------------

def strip_external(name: str) -> tuple:
    """Remove a trailing [External]-style tag; return (clean_name, is_external)."""
    n = name or ""
    low = n.lower()
    ext = "[external" in low or "(external" in low
    for tag in ("[external]", "(external)", "[external ]", "- external"):
        idx = low.find(tag)
        if idx != -1:
            n = n[:idx]
            break
    return n.strip(" -"), ext


def norm_name(name: str) -> str:
    clean, _ = strip_external(name)
    return " ".join(clean.lower().replace(".", " ").replace(",", " , ").split())


def parse_name_variants(name: str) -> list:
    """-> list of (surname, [given tokens]) candidate parses, most likely first.
    A comma is unambiguous: 'Surname, Given...'. Without one we can't tell
    which side is the surname - CMDB's 'Assigned to'/'Owner' writes
    'Given [Middle] Surname' (e.g. 'Michele De Filippo', surname = 'De
    Filippo'), so try every trailing chunk as a (possibly multi-word)
    surname, shortest first, plus the legacy single-word-prefix reading as a
    last resort. Caller tries each until one resolves against AD."""
    clean, _ = strip_external(name)
    clean = clean.lower()
    if "," in clean:
        surname, given = clean.split(",", 1)
        given_tokens = [t for t in given.replace("-", " ").replace(".", " ").split() if t]
        return [(surname.strip(), given_tokens)]

    parts = [t for t in clean.replace("-", " ").replace(".", " ").split() if t]
    if len(parts) <= 1:
        return [(parts[0] if parts else "", [])]

    variants = [(" ".join(parts[len(parts) - k:]), tuple(parts[:len(parts) - k]))
                for k in range(1, len(parts))]
    variants.append((parts[0], tuple(parts[1:])))  # legacy 'Surname Given...' fallback
    seen, out = set(), []
    for surname, given_tuple in variants:
        v = (surname, given_tuple)
        if v not in seen:
            seen.add(v)
            out.append((surname, list(given_tuple)))
    return out


def parse_name(name: str) -> tuple:
    """-> (surname, [given tokens]), single best-guess parse. See
    parse_name_variants() for callers that need to try multiple readings."""
    return parse_name_variants(name)[0]


def name_tokens(text: str) -> set:
    """Lowercase, split on space/hyphen/period/comma -> set of tokens.
    Used on structured GivenName values (e.g. 'Terry' or 'Tai Man')."""
    t = (text or "").lower().replace("-", " ").replace(".", " ").replace(",", " ")
    return {tok for tok in t.split() if tok}


def read_ad_users() -> dict:
    """Build lookup structures from the AD_Users export. Prefers the authoritative
    Surname/GivenName columns when present; falls back to parsing DisplayName
    only for rows where those columns are blank, so it degrades gracefully if
    a real export happens not to carry them."""
    found = find_dataset(AD_USERS["stem"])
    if not found:
        print("No AD_Users file/sheet in data/ - names can't be resolved to emails.")
        return {"exact": {}, "by_surname": {}, "count": 0}
    path, sheet = found
    m = AD_USERS["map"]
    exact, by_surname = {}, {}
    n = n_structured = 0
    for raw in _read_any(path, sheet):
        disp = cell_to_str(raw.get(m["name"], ""))
        email = cell_to_str(raw.get(m["email"], ""))
        if not disp or "@" not in email:
            continue
        n += 1
        exact[norm_name(disp)] = email

        surname_col = cell_to_str(raw.get(m.get("surname", ""), ""))
        given_col = cell_to_str(raw.get(m.get("given", ""), ""))
        if surname_col:
            n_structured += 1
            surname = surname_col.strip().lower()
            given_tokens = name_tokens(given_col)
        else:
            surname, given_list = parse_name(disp)
            given_tokens = set(given_list)

        by_surname.setdefault(surname, []).append(
            {"disp": disp, "email": email, "given": given_tokens})
    where = f"{os.path.basename(path)}" + (f" [{sheet}]" if sheet else "")
    print(f"AD_Users '{where}': {n} name->email entries "
          f"({n_structured} using Surname/GivenName columns, "
          f"{n - n_structured} parsed from DisplayName).")
    return {"exact": exact, "by_surname": by_surname, "count": n}


def read_overrides() -> dict:
    found = find_dataset(OVERRIDES["stem"])
    if not found:
        return {}
    path, sheet = found
    m = OVERRIDES["map"]
    out = {}
    for raw in _read_any(path, sheet):
        name = cell_to_str(raw.get(m["name"], ""))
        email = cell_to_str(raw.get(m["email"], ""))
        if name and "@" in email:
            out[norm_name(name)] = email
    if out:
        where = f"{os.path.basename(path)}" + (f" [{sheet}]" if sheet else "")
        print(f"Overrides '{where}': {len(out)} manual entries.")
    return out


def resolve_name_to_email(name: str, ad: dict, overrides: dict) -> tuple:
    """-> (email_or_None, method, confidence, candidate_emails).
    Confidence: high (override/exact), medium (unique heuristic), low (review)."""
    key = norm_name(name)
    if key in overrides:
        return overrides[key], "override", "high", []
    if key in ad["exact"]:
        return ad["exact"][key], "exact name", "high", []

    # A comma-less name is ambiguous about which part is the surname (see
    # parse_name_variants), so try every plausible reading and return on the
    # first one that resolves cleanly. Remember the first reading that at
    # least found a surname bucket, so an ambiguous-but-real match still
    # goes to review instead of being reported as "no AD match" just because
    # a *later*, wrong reading found nothing.
    fallback = None
    for surname, given_list in parse_name_variants(name):
        given = set(given_list)
        candidates = ad["by_surname"].get(surname, [])
        if not candidates:
            continue
        # candidates sharing at least one given-name token (e.g. "terry").
        # A shared *count* of 1 (just "terry") is weaker evidence than 2
        # ("terry" + "sp"), so rank by overlap size instead of treating any
        # overlap as equally good - e.g. query 'Terry-SP Lau' -> {terry, sp}
        # should prefer AD's 'Terry-SP.Lau' ({terry, sp}, full match) over
        # 'Terry-CP.Lau' ({terry}, partial match), not flag both as tied.
        shared = [c for c in candidates if given & c["given"]]
        if shared:
            best = max(len(given & c["given"]) for c in shared)
            shared = [c for c in shared if len(given & c["given"]) == best]
        if len(shared) == 1:
            return shared[0]["email"], "heuristic (surname+given)", "medium", []
        if fallback is not None:
            continue
        if len(shared) > 1:
            fallback = (None, "review: several AD names share surname+given", "low",
                        [c["email"] for c in shared])
        elif len(candidates) == 1:
            fallback = (None, "review: surname-only match (no given overlap)", "low",
                        [candidates[0]["email"]])
        else:
            fallback = (None, "review: surname matches several, no given overlap", "low",
                        [c["email"] for c in candidates])
    if fallback:
        return fallback
    return None, "no AD match (surname)", "none", []


def resolve_recipient(row: dict, cmdb_names: dict, ad: dict, overrides: dict) -> tuple:
    """-> (email_or_None, how, confidence, candidate_emails). Only ever called
    for workstation rows - build_notifications() skips servers before this."""
    at = (row.get("assigned_to") or "").strip()
    if "@" in at:                                   # Purview sometimes has the email directly
        return at, "user (assigned_to email)", "high", []

    name = at or cmdb_names.get((row.get("hostname") or "").strip().upper(), "")
    if not name:
        return None, "unresolved: no assigned user (CMDB/assigned_to)", "none", []

    email, method, conf, cands = resolve_name_to_email(name, ad, overrides)
    if email and conf in NOTIFY_CONFIDENCE:
        return email, f"user ({method})", conf, []
    return None, f"{method}: '{name}'", conf, cands


# ===========================================================================
# MESSAGE COMPOSITION  (one consolidated message per recipient)
# ===========================================================================

def compose_email(findings: list) -> tuple:
    """User-facing message: names the affected device(s) only, no per-source
    technical detail (issue/action text, CMDB upkeep) - that stays in the
    Worklist for staff. See USER_FACING_ACTION."""
    hosts = sorted({f["hostname"] for f in findings})

    subject = f"Action required: {len(hosts)} device(s) need a security update"
    lines = ["Hello,", "",
             f"The following device(s) associated with you are currently flagged "
             f"non-compliant and need attention as soon as possible:", ""]
    for host in hosts:
        lines.append(f"* {host}")
    lines += ["", USER_FACING_ACTION, "",
              "If you continue to see this notice after doing so, please contact IT Support.", "",
              "Thank you,", FROM_TEAM]
    return subject, "\n".join(lines)


def compose_email_html(findings: list) -> str:
    """HTML counterpart to compose_email() - presentation only, same content:
    same greeting, same host list, same urgency line, same USER_FACING_ACTION,
    same sign-off. Deliberately does NOT reintroduce the per-finding
    issue/action detail compose_email() already dropped for end users - see
    that function's docstring. compose_email() remains the plain-text
    version (used for the subject line and the dry-run preview); this is an
    additional HTML rendering, not a replacement.

    Every interpolated value is html.escape()'d before being placed into the
    markup - hostnames come from the source reports and must never be
    trusted as safe-to-embed raw text. Simple inline-styled markup only (no
    external CSS/fonts/JS/remote images) so it renders predictably in a
    conservative client like Outlook.
    """
    hosts = sorted({f["hostname"] for f in findings})

    host_items = "".join(
        f'<li style="font-size:16px; font-weight:bold; margin:6px 0; color:#1a1a1a;">'
        f'{html.escape(host)}</li>'
        for host in hosts
    )

    return f"""<div style="font-family:'Segoe UI',Arial,sans-serif; font-size:14px; color:#1a1a1a; line-height:1.5;">
<p>Hello,</p>
<p>The following device(s) associated with you are currently flagged non-compliant and need attention as soon as possible:</p>
<ul style="margin:8px 0 16px 20px; padding:0;">
{host_items}
</ul>
<p>{html.escape(USER_FACING_ACTION)}</p>
<p>If you continue to see this notice after doing so, please contact IT Support.</p>
<p>Thank you,<br>{html.escape(FROM_TEAM)}</p>
</div>"""


def compose_teams(findings: list) -> str:
    hosts = sorted({f["hostname"] for f in findings})
    return (f"You have {len(hosts)} non-compliant device(s) needing attention as soon as "
            f"possible: {', '.join(hosts)}. {USER_FACING_ACTION}")


def build_notifications(rows: list, cmdb_names: dict, ad: dict, overrides: dict) -> tuple:
    groups, review, unresolved = {}, [], []
    for r in rows:
        if r["kind"] == "Server":
            continue  # servers have no end user to notify - tracked in the worklist only
        email, how, conf, cands = resolve_recipient(r, cmdb_names, ad, overrides)
        if email:
            g = groups.setdefault(email, {"how": how, "rows": []})
            g["rows"].append(r)
        elif conf == "low":                       # has a name, but uncertain -> review
            review.append((r, how, cands))
        else:                                     # no user / no team at all
            unresolved.append((r, how))
    return groups, review, unresolved


def write_notifications_preview(groups: dict, review: list, unresolved: list) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if not HAVE_XLSX:
        path = os.path.join(OUTPUT_DIR, "notifications_preview.csv")
        with open(path, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["Recipient", "ResolvedBy", "Hosts", "Findings", "EmailSubject", "EmailBody", "TeamsMessage"])
            for email, g in sorted(groups.items()):
                subj, body = compose_email(g["rows"])
                hosts = len({r["hostname"] for r in g["rows"]})
                w.writerow([email, g["how"], hosts, len(g["rows"]), subj, body, compose_teams(g["rows"])])
        return path

    path = os.path.join(OUTPUT_DIR, "notifications_preview.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notifications"
    ws.append(["Recipient", "Resolved By", "Hosts", "Findings", "Email Subject", "Email Body", "Teams Message"])
    for email, g in sorted(groups.items()):
        subj, body = compose_email(g["rows"])
        hosts = len({r["hostname"] for r in g["rows"]})
        ws.append([email, g["how"], hosts, len(g["rows"]), subj, body, compose_teams(g["rows"])])
    ws.freeze_panes = "A2"

    # Review: has a name, but the match was uncertain -> a human picks the right
    # email and adds it to the overrides file. NOTHING here is emailed.
    rv = wb.create_sheet("Review")
    rv.append(["Hostname", "Source", "Kind", "BU", "Why held", "Possible emails (pick one -> overrides)"])
    for r, how, cands in review:
        rv.append([r["hostname"], r["source"], r["kind"], r["bu"], how, "  |  ".join(cands)])
    rv.freeze_panes = "A2"

    ur = wb.create_sheet("Unresolved")
    ur.append(["Hostname", "Source", "Platform", "Kind", "BU", "Reason no recipient"])
    for r, how in unresolved:
        ur.append([r["hostname"], r["source"], r["platform"], r["kind"], r["bu"], how])
    ur.freeze_panes = "A2"
    wb.save(path)
    return path


def write_html_preview(groups: dict) -> str:
    """Renders every recipient's compose_email_html() output into ONE combined
    HTML file under OUTPUT_DIR, so the actual rendered formatting can be
    opened in a browser and eyeballed before any real send - a pure review
    aid, nothing here is sent. Only the confidently-resolved `groups` are
    rendered (same set write_notifications_preview()'s Notifications sheet
    covers) - review/unresolved findings are never emailed, so there's
    nothing to preview a rendering of for them.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, "notifications_preview.html")

    sections = []
    for email, g in sorted(groups.items()):
        subject, _ = compose_email(g["rows"])
        body_html = compose_email_html(g["rows"])
        sections.append(f"""
<section style="border:1px solid #ccc; border-radius:6px; margin:0 0 24px 0; padding:16px; max-width:700px; background:#fff;">
  <div style="font-family:'Segoe UI',Arial,sans-serif; font-size:12px; color:#666; margin-bottom:8px;">
    <strong>To:</strong> {html.escape(email)}<br>
    <strong>Subject:</strong> {html.escape(subject)}
  </div>
  <hr style="border:none; border-top:1px solid #eee; margin:8px 0 16px 0;">
  {body_html}
</section>""")

    page = f"""<!doctype html>
<html>
<head><meta charset="utf-8"><title>Notification preview - {len(groups)} recipient(s)</title></head>
<body style="font-family:'Segoe UI',Arial,sans-serif; background:#f4f4f4; padding:24px; margin:0;">
<h2 style="font-family:'Segoe UI',Arial,sans-serif;">Notification preview - {len(groups)} recipient(s) - NOTHING SENT</h2>
{"".join(sections)}
</body>
</html>
"""
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(page)
    return path


def print_notify_summary(groups: dict, review: list, unresolved: list) -> None:
    by_how = {}
    for g in groups.values():
        by_how[g["how"]] = by_how.get(g["how"], 0) + 1
    print("\n" + "#" * 72)
    print(f"# {len(groups)} recipient(s) to notify   |   "
          f"{len(review)} finding(s) HELD FOR REVIEW   |   {len(unresolved)} UNRESOLVED")
    print("#" * 72)
    for how, n in sorted(by_how.items()):
        print(f"  {n:3d} recipient(s) via {how}")
    if review:
        print("\n  HELD FOR REVIEW (uncertain name match - not emailed):")
        for r, how, cands in review:
            print(f"    - {r['hostname']:16s} {how}")
    if unresolved:
        print("\n  UNRESOLVED (no user/team):")
        for r, how in unresolved:
            print(f"    - {r['hostname']:16s} {r['kind']:11s} {r['bu']:12s} [{how}]")


# ===========================================================================
# MOCK DATA  (real schemas; doubles as a schema fixture / test)
# ===========================================================================

def _write_mock(name: str, headers: list, rows: list) -> None:
    if HAVE_XLSX:
        p = os.path.join(DATA_DIR, f"{name}.xlsx")
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        wb.save(p)
    else:
        p = os.path.join(DATA_DIR, f"{name}.csv")
        with open(p, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=headers); w.writeheader(); w.writerows(rows)
    print(f"  wrote {p}")


def generate_mock_data() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    print(f"Generating mock data in '{DATA_DIR}/' ({'xlsx' if HAVE_XLSX else 'csv'}):")
    RD = "2026-06-30"

    # Shapes taken directly from the real file's actual schema/values (see
    # the "crowdstrike" FILE_REGISTRY comment) - a simple binary compliant/
    # non-compliant export, no separate "outdated" reason field.
    _write_mock("Crowdstrike_Deployment", FILE_REGISTRY["crowdstrike"]["columns"],
        [
            {"business_unit_code": "APAC-Retail", "host_name": "WS-APAC-001", "install_status": "Installed",
             "os": "Windows 11 Enterprise", "assigned_to": "Chan, Tai Man Terry",
             "crowdstrike": "No", "crowdstrike_installed": "0", "compliant": "0",
             "agent_version": "", "last_seen": "", "report_date": RD},
            # assigned_to blank -> falls back to CMDB hostname->name lookup
            {"business_unit_code": "APAC-Retail", "host_name": "WS-APAC-002", "install_status": "Installed",
             "os": "Windows 11 Enterprise", "assigned_to": "",
             "crowdstrike": "No", "crowdstrike_installed": "0", "compliant": "0",
             "agent_version": "", "last_seen": "", "report_date": RD},
            # assigned_to blank; CMDB's name for this host (Kelvin Lam) is
            # given-name-first with no comma - only resolves correctly via
            # AD's Surname/GivenName columns, not by parsing DisplayName
            {"business_unit_code": "EMEA-Ops", "host_name": "WS-EMEA-014", "install_status": "Installed",
             "os": "Windows 11 Enterprise", "assigned_to": "",
             "crowdstrike": "No", "crowdstrike_installed": "0", "compliant": "0",
             "agent_version": "7.36.20805.0", "last_seen": "2026-06-28", "report_date": RD},
            # assigned_to blank; CMDB carries the full 'Terry-SP' suffix - two
            # AD names share the given token "terry" ('Terry' vs 'Terry-SP'),
            # must resolve uniquely to the fuller-overlap candidate, not review
            {"business_unit_code": "APAC-Retail", "host_name": "WS-APAC-005", "install_status": "Installed",
             "os": "Windows 11 Enterprise", "assigned_to": "",
             "crowdstrike": "No", "crowdstrike_installed": "0", "compliant": "0",
             "agent_version": "", "last_seen": "2026-06-15", "report_date": RD},
            # compliant device in this UNFILTERED export -> must be skipped
            {"business_unit_code": "APAC-Retail", "host_name": "WS-APAC-011", "install_status": "Installed",
             "os": "Windows 11 Enterprise", "assigned_to": "Wong, Siu Ming",
             "crowdstrike": "Yes", "crowdstrike_installed": "1", "compliant": "1",
             "agent_version": "7.36.20805.0", "last_seen": "2026-06-29", "report_date": RD},
        ])

    # Real Mac CrowdStrike export shape (see "(mac)-cs" FILE_REGISTRY entry) -
    # no assigned_to column at all, so recipient resolution relies entirely
    # on CMDB fallback. Revives MAC-APAC-07/MAC-AMS-22 as real findings again
    # (CMDB/AD_Users mock rows below already model these two: an [External]-
    # tagged name and an ambiguous "Smith, Robert" match) - both existed only
    # as inert fixture data after the old aiago_mac_cs entry (which used to
    # source them) was removed for relying on a file that no longer exists.
    _write_mock("Workstation_Deployment (MAC)-CS", FILE_REGISTRY["(mac)-cs"]["columns"],
        [
            {"hostname": "MAC-APAC-07", "business_unit_code": "APAC-Retail",
             "operating_system": "MAC", "compliant_status": "0",
             "cs_agent_version": "", "cs_last_seen": "", "report_date": RD},
            # ambiguous name (Smith, Robert) -> must land in review, not groups
            {"hostname": "MAC-AMS-22", "business_unit_code": "AMS-Corp",
             "operating_system": "MAC", "compliant_status": "0",
             "cs_agent_version": "", "cs_last_seen": "", "report_date": RD},
            # compliant device in this UNFILTERED export -> must be skipped
            {"hostname": "MAC-APAC-08", "business_unit_code": "APAC-Retail",
             "operating_system": "MAC", "compliant_status": "1",
             "cs_agent_version": "7.36.20807.0", "cs_last_seen": "2026-06-29", "report_date": RD},
        ])

    _write_mock("Workstation_Deployment-Zapp", FILE_REGISTRY["deployment-zapp"]["columns"],
        [
            # not compliant, client missing entirely -> a finding
            {"hostname": "WS-APAC-001", "business_unit_code": "APAC-Retail", "zapp": "FALSE",
             "assigned_to": "Terry Chan", "install_status": "Installed", "os": "Windows 11 24H2",
             "sys_class_name": "Computer", "zapp_installed": "0", "zapp_missing": "1",
             "compliant": "0", "report_date": RD},
            # compliant device in this UNFILTERED export -> must be skipped, not emailed
            {"hostname": "WS-EMEA-030", "business_unit_code": "EMEA-Ops", "zapp": "TRUE",
             "assigned_to": "carol@example.com", "install_status": "Installed", "os": "Windows 11 24H2",
             "sys_class_name": "Computer", "zapp_installed": "1", "zapp_missing": "0",
             "compliant": "1", "report_date": RD},
        ])

    # Real Mac Zapp export shape (see "(mac)-zapp" FILE_REGISTRY entry) - also
    # no assigned_to column, so this host needs its own CMDB entry (below).
    _write_mock("Workstation_Deployment (MAC)-Zapp", FILE_REGISTRY["(mac)-zapp"]["columns"],
        [
            {"hostname": "MAC-AMS-25", "business_unit_code": "AMS-Corp",
             "operating_system": "MAC", "compliant_status": "0",
             "zapp_version": "", "last_seen_connected_to_zia": "", "report_date": RD},
            # compliant device in this UNFILTERED export -> must be skipped
            {"hostname": "MAC-APAC-09", "business_unit_code": "APAC-Retail",
             "operating_system": "MAC", "compliant_status": "1",
             "zapp_version": "4.5.2.105", "last_seen_connected_to_zia": "2026-06-29", "report_date": RD},
        ])

    _write_mock("Workstation_Deployment-DLP", FILE_REGISTRY["deployment-dlp"]["columns"],
        [
            # same host as the Crowdstrike mock's WS-APAC-001 above, but a
            # DIFFERENT source (Purview, not CrowdStrike) - must consolidate
            # into ONE email for Terry (one message per person, not one per
            # source), not be deduped away (dedup is keyed on (hostname,
            # source), so these two rows are never considered duplicates).
            {"name": "WS-APAC-001", "business_unit_code": "APAC-Retail", "assigned_to": "Chan, Tai Man Terry",
             "install_status": "Installed", "os": "Windows 11 24H2", "sys_class_name": "Computer",
             "purview_configuration_status": "NotUpdated", "purview_policy_status": "NotUpdated",
             "purview_last_seen": "2026-06-25", "compliance": "Non-compliant", "report_date": RD},
            # a host with a Purview finding but no CrowdStrike one - DLP is
            # now the sole Windows Purview-sourced entry, so there's nothing
            # for it to be "added on top of"; this just pins that a host
            # appearing in only one source still becomes a finding
            {"name": "WS-APAC-006", "business_unit_code": "APAC-Retail", "assigned_to": "Wong, Siu Ming",
             "install_status": "Installed", "os": "Windows 11 24H2", "sys_class_name": "Computer",
             "purview_configuration_status": "", "purview_policy_status": "",
             "purview_last_seen": "", "compliance": "Non-compliant", "report_date": RD},
            # compliant device in this UNFILTERED export -> must be skipped
            {"name": "WS-APAC-007", "business_unit_code": "APAC-Retail", "assigned_to": "carol@example.com",
             "install_status": "Installed", "os": "Windows 11 24H2", "sys_class_name": "Computer",
             "purview_configuration_status": "Updated", "purview_policy_status": "Updated",
             "purview_last_seen": "2026-06-29", "compliance": "Compliant", "report_date": RD},
        ])

    # Real Mac DLP/Purview export shape (see "(mac)-dlp" FILE_REGISTRY entry).
    # MAC-AMS-22 reused here too - a second (third, counting Zapp... no,
    # CS+DLP) finding for the same ambiguous "Smith, Robert" host, same
    # pattern as WS-APAC-001's multi-source consolidation above.
    _write_mock("Workstation_Deployment (MAC)-DLP", FILE_REGISTRY["(mac)-dlp"]["columns"],
        [
            {"intune_computer_name": "MAC-AMS-22", "business_unit_code": "AMS-Corp",
             "purview_configuration_status": "", "purview_policy_status": "",
             "purview_last_seen": "", "compliance": "Non-compliant", "report_date": RD},
            # compliant device in this UNFILTERED export -> must be skipped
            # (matches the real snapshot: 42/42 rows Compliant today)
            {"intune_computer_name": "MAC-APAC-10", "business_unit_code": "APAC-Retail",
             "purview_configuration_status": "Updated", "purview_policy_status": "Updated",
             "purview_last_seen": "2026-06-29", "compliance": "Compliant", "report_date": RD},
        ])

    # BitLocker (Hard Disk Encryption Compliance) export - shapes taken
    # directly from the real file's actual non-compliant rows (see
    # bitlocker_issue()'s docstring for the full breakdown).
    _write_mock("Hard_Disk_Encryption_Compliance", FILE_REGISTRY["encryption"]["columns"],
        [
            # compliant device in this UNFILTERED export -> must be skipped
            {"host_name": "WS-APAC-008", "business_unit_code": "APAC-Retail", "assigned_to": "Wong, Siu Ming",
             "install_status": "Installed", "os": "Windows 11 Enterprise", "encryption_status": "encrypted",
             "setting_state_summary": "compliant", "compliant": "Compliant",
             "last_discovered": "2026-07-10 12:00:00", "report_date": RD},
            # the clearest real-world non-compliant shape: reports in, drive
            # explicitly not encrypted
            {"host_name": "WS-APAC-009", "business_unit_code": "APAC-Retail", "assigned_to": "Wong, Siu Ming",
             "install_status": "Installed", "os": "Windows 11 Enterprise", "encryption_status": "notEncrypted",
             "setting_state_summary": "compliant", "compliant": "Non-compliant",
             "last_discovered": "2026-06-15 09:30:00", "report_date": RD},
            # the majority real-world non-compliant shape: no BitLocker/Intune
            # telemetry reported at all
            {"host_name": "WS-APAC-010", "business_unit_code": "APAC-Retail", "assigned_to": "Chan, Tai Man Terry",
             "install_status": "Installed", "os": "Windows 11 Enterprise", "encryption_status": "",
             "setting_state_summary": "", "compliant": "Non-compliant",
             "last_discovered": "", "report_date": RD},
        ])

    # CMDB export: hostname ('Name') -> assigned user DISPLAY NAME ('Assigned to').
    # Names use the fuller convention; AD_Users below uses the shorter one.
    # WS-EMEA-014 absent -> stays unresolved.
    _write_mock("CMDB_Mapping",
        ["Name", "Serial number", "Assigned to", "Install Status", "Operating System"],
        [
            {"Name": "WS-APAC-001", "Serial number": "SN-A1", "Assigned to": "Chan, Tai Man Terry", "Install Status": "Installed", "Operating System": "Windows 11 24H2"},
            {"Name": "WS-APAC-002", "Serial number": "SN-A2", "Assigned to": "Wong, Siu Ming", "Install Status": "Installed", "Operating System": "Windows 11 24H2"},
            {"Name": "MAC-APAC-07", "Serial number": "SN-A7", "Assigned to": "Lee, John Xavier [External]", "Install Status": "Installed", "Operating System": "macOS 14.5"},
            {"Name": "MAC-AMS-22", "Serial number": "SN-M22", "Assigned to": "Smith, Robert", "Install Status": "Installed", "Operating System": "macOS 14.4"},
            {"Name": "WS-EMEA-014", "Serial number": "SN-E14", "Assigned to": "Lam, Wai Lok Kelvin", "Install Status": "Installed", "Operating System": "Windows 11 24H2"},
            {"Name": "WS-APAC-005", "Serial number": "SN-A5", "Assigned to": "Terry-SP Lau", "Install Status": "Installed", "Operating System": "Windows 11 24H2"},
            {"Name": "MAC-AMS-25", "Serial number": "SN-M25", "Assigned to": "Wong, Siu Ming", "Install Status": "Installed", "Operating System": "macOS 14.4"},
        ])

    # AD directory: DisplayName -> EmailAddress. Note the shorter convention and
    # the two "Smith, Robert-*" rows that make "Smith, Robert" ambiguous.
    # Real AD exports often carry Surname/GivenName as separate columns rather
    # than relying on DisplayName formatting. Kelvin Lam demonstrates exactly
    # why that matters: his DisplayName is given-name-first with NO comma, so
    # naive parsing would misread "Kelvin" as the surname. The Surname/GivenName
    # columns sidestep that entirely.
    _write_mock("AD_Users",
        ["DisplayName", "Surname", "GivenName", "EmailAddress", "Department"],
        [
            {"DisplayName": "Chan, Terry-TM", "Surname": "Chan", "GivenName": "Terry", "EmailAddress": "terry.chan@example.com", "Department": "Retail"},
            {"DisplayName": "Wong, Siu Ming", "Surname": "Wong", "GivenName": "Siu Ming", "EmailAddress": "siuming.wong@example.com", "Department": "Ops"},
            {"DisplayName": "Lee, John-JX [External]", "Surname": "Lee", "GivenName": "John", "EmailAddress": "john.lee@consultant.com", "Department": "Contractor"},
            {"DisplayName": "Smith, Robert-RA", "Surname": "Smith", "GivenName": "Robert", "EmailAddress": "robert.a.smith@example.com", "Department": "Finance"},
            {"DisplayName": "Smith, Robert-RB", "Surname": "Smith", "GivenName": "Robert", "EmailAddress": "robert.b.smith@example.com", "Department": "Legal"},
            {"DisplayName": "Kelvin Lam", "Surname": "Lam", "GivenName": "Kelvin", "EmailAddress": "kelvin.lam@example.com", "Department": "IT"},
            # 'Lau' shares the given token "terry" between both rows, but only
            # Terry-SP's given carries the full {"terry","sp"} that matches the
            # CMDB name above - regression test for the overlap-size ranking.
            {"DisplayName": "Lau, Terry-CP", "Surname": "Lau", "GivenName": "Terry", "EmailAddress": "terry-cp.lau@example.com", "Department": "Retail"},
            {"DisplayName": "Lau, Terry-SP", "Surname": "Lau", "GivenName": "Terry-SP", "EmailAddress": "terry-sp.lau@example.com", "Department": "Legal"},
        ])


# ===========================================================================
# PRE-FLIGHT HEADER CHECK
# Reads each file's real header row and reports, by name, any column the code
# expects but can't find - so a renamed/mismatched column announces itself
# instead of silently producing blank fields. Warns; never blocks the run.
# ===========================================================================

def validate_headers() -> bool:
    print("Pre-flight header check:")
    all_ok = True

    checks = []  # (stem_keyword, required_columns, optional_columns, label, expected_headers)
    for rk, entry in FILE_REGISTRY.items():
        checks.append((rk, list(entry["map"].values()), [],
                        f"{entry['meta']['source']}/{entry['meta']['kind']}", entry.get("columns")))
    checks.append((CMDB_MAPPING["stem"], list(CMDB_MAPPING["map"].values()), [], "CMDB", None))
    checks.append((AD_USERS["stem"], [AD_USERS["map"]["name"], AD_USERS["map"]["email"]],
                   [AD_USERS["map"].get("surname", ""), AD_USERS["map"].get("given", "")], "AD_Users", None))
    checks.append((OVERRIDES["stem"], list(OVERRIDES["map"].values()), [], "Overrides", None))

    for stem_keyword, required, optional, label, expected_headers in checks:
        found = find_dataset(stem_keyword)
        if not found:
            if label == "Overrides":
                continue  # optional file, silently skip if absent
            all_ok = False
            print(f"  !  {label}: no matching file or sheet found in '{DATA_DIR}/'")
            continue
        path, sheet = found
        actual = _read_headers(path, sheet, expected_headers)
        where = f"{os.path.basename(path)}" + (f" [{sheet}]" if sheet else "")
        missing = [c for c in required if c not in actual]
        missing_opt = [c for c in optional if c and c not in actual]
        if missing:
            all_ok = False
            print(f"  !  {where}  ({label}) missing column(s): {', '.join(missing)}")
            print(f"       actually has: {', '.join(actual) or '(no headers)'}")
        else:
            note = f"  (no {'/'.join(missing_opt)} - will parse names from DisplayName instead)" if missing_opt else ""
            print(f"  OK {where}  ({label}){note}")

    if not all_ok:
        print("  -> Update the column name(s) in FILE_REGISTRY / CMDB_MAPPING / AD_USERS")
        print("     to match the 'file actually has' list above, then re-run.")
    print()
    return all_ok


def _warn_if_thresholds_stale() -> bool:
    """CS_LATEST/PURVIEW_LATEST are hardcoded reference versions with no way
    to auto-refresh from here - the only defense against them silently going
    stale is making staleness itself visible. -> True if stale."""
    age_days = (date.today() - THRESHOLDS_VERIFIED).days
    if age_days > THRESHOLDS_STALE_AFTER_DAYS:
        print(f"  ! CS_LATEST/PURVIEW_LATEST were last verified {age_days} days ago "
              f"(on {THRESHOLDS_VERIFIED.isoformat()}) - CrowdStrike/Purview have "
              f"likely shipped newer clients since. Check the current release "
              f"versions and update CS_LATEST/PURVIEW_LATEST (and THRESHOLDS_VERIFIED) "
              f"near the top of this file.")
        print()
        return True
    return False


# ===========================================================================
# ENTRY POINT
# ===========================================================================

def _data_dir_has_files() -> bool:
    return any(f.lower().endswith((".xlsx", ".csv")) for f in _get_data_source().list_files())


def main() -> None:
    print(f"[i/o mode: {'XLSX' if HAVE_XLSX else 'CSV'}]\n")
    if "--regen" in sys.argv or not _data_dir_has_files():
        generate_mock_data()
        print()

    thresholds_stale = _warn_if_thresholds_stale()
    preflight_ok = validate_headers()

    rows = load_all()
    if not rows:
        print("No rows loaded - check that report files are in data/.")
        sys.exit(1)

    worklist = write_worklist(rows)

    cmdb_names = read_cmdb_mapping()
    ad = read_ad_users()
    overrides = read_overrides()
    groups, review, unresolved = build_notifications(rows, cmdb_names, ad, overrides)
    preview = write_notifications_preview(groups, review, unresolved)
    html_preview = write_html_preview(groups)
    print_notify_summary(groups, review, unresolved)

    print(f"\nConsolidated worklist   -> {worklist}")
    print(f"Notification preview    -> {preview}   (NOTHING SENT)")
    print(f"HTML preview (open in a browser) -> {html_preview}   (NOTHING SENT)")

    if not preflight_ok or thresholds_stale:
        print("\n" + "!" * 72)
        if not preflight_ok:
            print("! Pre-flight check found problem(s) (see 'Pre-flight header check' above) -")
            print("! this run's output may be missing data for the affected report(s).")
        if thresholds_stale:
            print("! CS_LATEST/PURVIEW_LATEST reference versions may be stale (see above).")
        print("!" * 72)
        sys.exit(1)


if __name__ == "__main__":
    main()